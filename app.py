"""
RemuPro v2.5 - Sistema de Procesamiento de Remuneraciones Educativas
Interfaz Web con Streamlit
"""

import re
import socket
import streamlit as st
import pandas as pd
from pathlib import Path
from io import BytesIO
import tempfile
from datetime import datetime

from processors import (
    SEPProcessor, PIEProcessor, DuplicadosProcessor,
    BRPProcessor, IntegradoProcessor, REMProcessor,
    EIBProcessor, AnualBatchProcessor,
)
from reports import AuditLog, InformeWord
from database import BRPRepository, ComparadorMeses
from config.columns import format_rut, normalize_rut, detect_month_from_filename, detect_file_type, detect_year_from_filename, MESES_NUM_TO_NAME
from config.escuelas import get_rbd_map, match_ubicacion
import html as html_module
import json
import streamlit.components.v1 as st_components


def _sanitize_html(text: str) -> str:
    """Sanitize text for safe HTML rendering, preserving allowed formatting tags.

    Escapes all HTML entities first, then re-enables a small whitelist of
    formatting tags (<b>, <br>) that the application uses intentionally.
    This prevents XSS from user-controlled data (file names, error messages)
    while keeping the existing UI formatting intact.
    """
    safe = html_module.escape(str(text))
    # Re-enable only the specific safe tags used by the application
    safe = safe.replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
    safe = safe.replace("&lt;br&gt;", "<br>").replace("&lt;br/&gt;", "<br/>")
    return safe


# ============================================================================
# CONFIGURACION
# ============================================================================

VERSION = "2.5.0"


def fmt_clp(value, prefix='$'):
    """Formatea un número al estilo chileno: $1.234.567 (puntos para miles)."""
    try:
        n = int(round(float(value)))
    except (ValueError, TypeError):
        return f"{prefix}0"
    negative = n < 0
    n = abs(n)
    formatted = f"{n:,}".replace(',', '.')
    sign = '-' if negative else ''
    return f"{sign}{prefix}{formatted}"


def fmt_clp_style(val):
    """Formateador para pandas style.format — formato chileno."""
    return fmt_clp(val)


def format_rbd(rbd):
    """Formatea un RBD con dígito verificador: 6710 → '6710-5'.

    Usa algoritmo módulo 11 (mismo que RUT chileno).
    """
    try:
        num = int(str(rbd).split('.')[0].split('-')[0].strip())
    except (ValueError, TypeError):
        return str(rbd)
    s = str(num)
    total = 0
    factor = 2
    for c in reversed(s):
        total += int(c) * factor
        factor = factor + 1 if factor < 7 else 2
    resto = 11 - (total % 11)
    if resto == 11:
        dv = '0'
    elif resto == 10:
        dv = 'K'
    else:
        dv = str(resto)
    return f"{num}-{dv}"


# Columnas numéricas que NO son dinero (horas, conteos, meses, etc.)
NON_MONEY_COLS = {
    'REGISTROS', 'DOCENTES', 'HORAS_SEP', 'HORAS_PIE', 'HORAS_SN',
    'HORAS_CONTRATO', 'MES_NUM', 'NUM_ESTABLECIMIENTOS', 'EXCESO',
    'MESES_PRESENTES', 'NUM_ESCUELAS', 'HORAS_TOTAL',
    'SEP', 'PIE', 'NORMAL', 'EIB', 'TOTAL', 'TOTAL HORAS',
    'Mes', 'SN',
}


def append_totals_row(df, label_col='MES', label='TOTAL', skip_cols=None, docentes_col=None, docentes_value=None):
    """Agrega fila TOTAL a un DataFrame.

    Args:
        df: DataFrame al cual agregar la fila.
        label_col: columna donde poner la etiqueta 'TOTAL'.
        label: texto de la etiqueta.
        skip_cols: set de columnas a dejar vacías.
        docentes_col: si se indica, pone docentes_value en vez de sumar.
        docentes_value: valor especial para la columna de docentes.
    """
    if df.empty:
        return df
    skip = set(skip_cols or [])
    totals = {}
    for col in df.columns:
        if col == label_col:
            totals[col] = label
        elif col in skip:
            totals[col] = ''
        elif docentes_col and col == docentes_col and docentes_value is not None:
            totals[col] = docentes_value
        elif df[col].dtype in ('float64', 'int64', 'float32', 'int32'):
            totals[col] = df[col].sum()
        else:
            totals[col] = ''
    return pd.concat([df, pd.DataFrame([totals])], ignore_index=True)


def format_money_cols(df, exclude_cols=None):
    """Retorna dict {col: fmt_clp_style} para columnas monetarias del DataFrame.

    Excluye automáticamente columnas no-monetarias (horas, conteos, etc.)
    y las indicadas en exclude_cols.
    """
    exclude = NON_MONEY_COLS.copy()
    if exclude_cols:
        exclude.update(exclude_cols)
    result = {}
    for col in df.columns:
        if col in exclude:
            continue
        if df[col].dtype in ('float64', 'int64', 'float32', 'int32'):
            result[col] = fmt_clp_style
    return result


def generate_pdf_from_df(df, title="Reporte", orientation='L', col_widths=None):
    """Genera PDF desde un DataFrame.

    Args:
        df: DataFrame a exportar.
        title: título del PDF.
        orientation: 'L' landscape o 'P' portrait.
        col_widths: dict {col: width_mm} opcional.
    Returns:
        bytes del PDF generado.
    """
    from fpdf import FPDF

    pdf = FPDF(orientation=orientation, unit='mm', format='A3')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Título
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, title, ln=True, align='C')
    pdf.ln(5)

    # Calcular anchos
    page_width = pdf.w - 2 * pdf.l_margin
    n_cols = len(df.columns)

    if col_widths:
        widths = [col_widths.get(col, page_width / n_cols) for col in df.columns]
    else:
        # Auto-calcular: columnas de texto más anchas, numéricas más estrechas
        widths = []
        for col in df.columns:
            if df[col].dtype == 'object':
                max_len = max(len(str(col)), df[col].astype(str).str.len().max() if not df.empty else 5)
                w = min(max(max_len * 2.2, 20), 60)
            else:
                w = max(len(str(col)) * 2.2, 18)
            widths.append(w)
        # Escalar para que quepan
        total = sum(widths)
        if total > page_width:
            scale = page_width / total
            widths = [w * scale for w in widths]

    # Header
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_fill_color(30, 41, 59)  # #1e293b
    pdf.set_text_color(255, 255, 255)
    for i, col in enumerate(df.columns):
        pdf.cell(widths[i], 7, str(col)[:20], border=1, align='C', fill=True)
    pdf.ln()

    # Body
    pdf.set_font('Helvetica', '', 6)
    pdf.set_text_color(0, 0, 0)

    for _, row in df.iterrows():
        is_total = False
        for col in df.columns:
            if str(row[col]).upper().startswith('TOTAL'):
                is_total = True
                break

        if is_total:
            pdf.set_font('Helvetica', 'B', 6)
            pdf.set_fill_color(241, 245, 249)  # #f1f5f9
        else:
            pdf.set_font('Helvetica', '', 6)
            pdf.set_fill_color(255, 255, 255)

        for i, col in enumerate(df.columns):
            val = row[col]
            if pd.notna(val) and col not in NON_MONEY_COLS and isinstance(val, (int, float)):
                display = fmt_clp(val)
            elif pd.notna(val):
                display = str(val)
            else:
                display = ''
            align = 'L' if df[col].dtype == 'object' else 'R'
            pdf.cell(widths[i], 5, display[:25], border=1, align=align, fill=is_total)
        pdf.ln()

    return pdf.output()


def parse_establishment_csv(file_content):
    """Parsea CSV de CPEIP por establecimiento con header multi-nivel.

    El CSV tiene estructura:
    - Columnas: RBD, ESTABLECIMIENTO, luego grupos de 3 por mes (Recon, Tramo, Prior)
    - Fila 0: sub-headers (nombres largos de conceptos)
    - Datos desde fila 1, última fila puede ser "Total general"

    Args:
        file_content: bytes o string path del CSV.
    Returns:
        DataFrame normalizado con formato {MES}_RECON, {MES}_TRAMO, {MES}_PRIOR.
    """
    if isinstance(file_content, bytes):
        df_raw = pd.read_csv(BytesIO(file_content))
    else:
        df_raw = pd.read_csv(file_content)

    if df_raw.empty or len(df_raw) < 2:
        return None

    # La fila 0 tiene los sub-headers reales, descartarla de los datos
    # Detectar si fila 0 es sub-header (contiene "Suma de" o "RBD" como valor en col RBD)
    first_val = str(df_raw.iloc[0, 0]).strip()
    if first_val.upper() == 'RBD' or 'suma de' in str(df_raw.iloc[0, 2]).lower():
        df_raw = df_raw.iloc[1:].reset_index(drop=True)

    # Reconstruir columnas: las named son meses, las Unnamed son sub-cols
    orig_cols = list(df_raw.columns)
    new_cols = []
    current_mes = None

    MONTH_MAP = {
        'ENERO': 'ENERO', 'FEBRERO': 'FEBRERO', 'MARZO': 'MARZO',
        'ABRIL': 'ABRIL', 'MAYO': 'MAYO', 'JUNIO': 'JUNIO',
        'JULIO': 'JULIO', 'AGOSTO': 'AGOSTO', 'SEPTIEMBRE': 'SEPTIEMBRE',
        'OCTUBRE': 'OCTUBRE', 'NOVIEMBRE': 'NOVIEMBRE', 'DICIEMBRE': 'DICIEMBRE',
    }
    suffixes = ['RECON', 'TRAMO', 'PRIOR']
    sub_idx = 0

    for col in orig_cols:
        col_clean = col.strip().upper()
        if col_clean in ('RBD',):
            new_cols.append('RBD')
            continue
        if 'ESTABLECIMIENTO' in col_clean:
            new_cols.append('ESTABLECIMIENTO')
            continue

        # Es un nombre de mes?
        matched_month = None
        for mkey in MONTH_MAP:
            if mkey in col_clean:
                matched_month = mkey
                break

        if matched_month:
            current_mes = matched_month
            sub_idx = 0
            new_cols.append(f'{current_mes}_{suffixes[sub_idx]}')
            sub_idx += 1
        elif col.startswith('Unnamed') and current_mes:
            if sub_idx < len(suffixes):
                new_cols.append(f'{current_mes}_{suffixes[sub_idx]}')
                sub_idx += 1
            else:
                new_cols.append(col)
        else:
            new_cols.append(col)

    df_raw.columns = new_cols

    # Limpiar valores: convertir strings con puntos como separador de miles
    for col in df_raw.columns:
        if col not in ('RBD', 'ESTABLECIMIENTO'):
            df_raw[col] = df_raw[col].apply(lambda x: _parse_clp_number(x))

    # Formatear RBD
    if 'RBD' in df_raw.columns:
        df_raw['RBD'] = df_raw['RBD'].apply(lambda x: format_rbd(str(x).strip()) if pd.notna(x) and not str(x).strip().upper().startswith('TOTAL') else str(x).strip() if pd.notna(x) else '')

    return df_raw


def _parse_clp_number(val):
    """Convierte string con formato CLP (1.234.567) a int."""
    if pd.isna(val):
        return 0
    s = str(val).strip()
    if s in ('-', '', 'nan', 'NaN'):
        return 0
    # Remover puntos de miles y posible $
    s = s.replace('$', '').replace(' ', '')
    # Si usa punto como separador de miles (no decimal)
    if '.' in s and ',' not in s:
        s = s.replace('.', '')
    elif ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


@st.cache_data
def _read_detalle_brp(excel_bytes):
    """Lee y cachea la hoja DETALLE_BRP del Excel."""
    try:
        return pd.read_excel(BytesIO(excel_bytes), sheet_name='DETALLE_BRP', engine='openpyxl')
    except Exception:
        return pd.DataFrame()


@st.cache_data
def _read_horas_completo(excel_bytes):
    """Lee y cachea la hoja HORAS_COMPLETO del Excel anual."""
    try:
        return pd.read_excel(BytesIO(excel_bytes), sheet_name='HORAS_COMPLETO', engine='openpyxl')
    except Exception:
        return pd.DataFrame()


@st.fragment
def _render_horas_tab(excel_bytes):
    """Renderiza la pestaña Horas del lote anual (HORAS_COMPLETO)."""
    st.markdown("### Horas por Subvención")
    st.caption("Horas de contrato clasificadas por tipo de subvención para cada docente y mes.")

    df = _read_horas_completo(excel_bytes)
    if df.empty:
        st.info("No hay datos de horas disponibles. Asegúrate de que el procesamiento incluya archivos de horas por subvención.")
        return

    # Filtros
    col_f1, col_f2 = st.columns([1, 2])
    with col_f1:
        meses = sorted(df['MES'].dropna().unique().tolist()) if 'MES' in df.columns else []
        mes_sel = st.selectbox("Filtrar por mes", ["Todos"] + meses, key="horas_tab_mes")
    with col_f2:
        busqueda = st.text_input("Buscar por nombre o RUT", key="horas_tab_buscar")

    df_f = df.copy()
    if mes_sel != "Todos":
        df_f = df_f[df_f['MES'] == mes_sel]

    if busqueda:
        busqueda_lower = busqueda.lower()
        mask = pd.Series(False, index=df_f.index)
        for col in df_f.columns:
            if col.lower() in ('rut', 'nombre'):
                mask = mask | df_f[col].astype(str).str.lower().str.contains(busqueda_lower, na=False)
        df_f = df_f[mask]

    # Métricas
    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        rut_col = _find_col(df_f, 'rut')
        n_docentes = df_f[rut_col].nunique() if rut_col else len(df_f)
        st.metric("Docentes", n_docentes)
    with mc2:
        total_col = _find_col(df_f, 'total')
        total_horas = int(df_f[total_col].sum()) if total_col and total_col in df_f.columns else 0
        st.metric("Total Horas", f"{total_horas:,}".replace(',', '.'))
    with mc3:
        n_registros = len(df_f)
        st.metric("Registros", n_registros)

    # Tabla
    df_display = append_totals_row(df_f, label_col='MES' if 'MES' in df_f.columns else df_f.columns[0], label='TOTAL')

    # Excluir MES_NUM de la vista
    cols_show = [c for c in df_display.columns if c != 'MES_NUM']
    st.dataframe(df_display[cols_show], width='stretch', hide_index=True)
    add_table_downloads(df_display[cols_show], 'horas_por_subvencion', 'horas_tab_dl')


def build_establishment_table(excel_bytes, summaries):
    """Construye tabla pivot CPEIP por establecimiento y mes.

    Retorna DataFrame con columnas multi-nivel: RBD, Escuela, y por cada mes
    (Reconocimiento, Tramo, Prioritarios).
    """
    df = _read_detalle_brp(excel_bytes)
    if df.empty:
        return None

    # Detectar columna RBD (puede ser 'RBD', 'Rbd (Establecimiento)', etc.)
    rbd_col = next((c for c in df.columns if 'rbd' in c.lower()), None)
    if not rbd_col:
        return None
    # Normalizar nombre a 'RBD' para el resto de la función
    if rbd_col != 'RBD':
        df = df.rename(columns={rbd_col: 'RBD'})

    # Columnas CPEIP totales por fila
    for concept, cols in [
        ('CPEIP_RECON', ['CPEIP_RECON_SEP', 'CPEIP_RECON_PIE', 'CPEIP_RECON_NORMAL']),
        ('CPEIP_TRAMO', ['CPEIP_TRAMO_SEP', 'CPEIP_TRAMO_PIE', 'CPEIP_TRAMO_NORMAL']),
        ('CPEIP_PRIOR', ['CPEIP_PRIOR_SEP', 'CPEIP_PRIOR_PIE', 'CPEIP_PRIOR_NORMAL']),
    ]:
        existing = [c for c in cols if c in df.columns]
        df[concept] = df[existing].sum(axis=1) if existing else 0

    # Detectar columna de mes
    mes_col = 'MES' if 'MES' in df.columns else None
    if not mes_col:
        for c in df.columns:
            if c.upper() in ('MES', 'PERIODO', 'MES_NUM'):
                mes_col = c
                break

    if not mes_col:
        return None

    # Lookup de nombres de escuela
    escuelas = get_rbd_map()

    # Agrupar por RBD y MES
    grouped = df.groupby(['RBD', mes_col]).agg({
        'CPEIP_RECON': 'sum',
        'CPEIP_TRAMO': 'sum',
        'CPEIP_PRIOR': 'sum',
    }).reset_index()

    # Obtener lista ordenada de meses
    all_months_order = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
    mes_map_to_num = {}
    for _, row in grouped.iterrows():
        m = str(row[mes_col]).strip()
        if m in MESES_NUM_TO_NAME:
            mes_map_to_num[m] = m
        else:
            for num, name in MESES_NUM_TO_NAME.items():
                if name.lower() == m.lower():
                    mes_map_to_num[m] = num
                    break

    meses_presentes = sorted(set(mes_map_to_num.values()), key=lambda x: all_months_order.index(x) if x in all_months_order else 99)

    # Construir tabla plana
    rbds = sorted(grouped['RBD'].unique(), key=lambda x: int(str(x).split('.')[0].split('-')[0]) if str(x).replace('.', '').replace('-', '').isdigit() else 999999)

    rows = []
    for rbd in rbds:
        rbd_str = str(rbd).split('.')[0].split('-')[0].strip()
        rbd_formatted = format_rbd(rbd_str)
        escuela = escuelas.get(rbd_str, '') if escuelas else ''
        row = {'RBD': rbd_formatted, 'ESTABLECIMIENTO': escuela}

        for mes_num in meses_presentes:
            mes_name = MESES_NUM_TO_NAME.get(mes_num, mes_num).upper()
            # Buscar el valor original del mes en grouped
            mask = grouped['RBD'] == rbd
            for orig_mes, mapped_num in mes_map_to_num.items():
                if mapped_num == mes_num:
                    mask_mes = mask & (grouped[mes_col] == orig_mes)
                    break
            else:
                mask_mes = mask & (grouped[mes_col] == mes_num)

            sub = grouped[mask_mes]
            recon = int(sub['CPEIP_RECON'].sum()) if not sub.empty else 0
            tramo = int(sub['CPEIP_TRAMO'].sum()) if not sub.empty else 0
            prior = int(sub['CPEIP_PRIOR'].sum()) if not sub.empty else 0

            row[f'{mes_name}_RECON'] = recon
            row[f'{mes_name}_TRAMO'] = tramo
            row[f'{mes_name}_PRIOR'] = prior

        rows.append(row)

    df_result = pd.DataFrame(rows)

    # Fila de totales
    totals = {'RBD': 'Total general', 'ESTABLECIMIENTO': ''}
    for col in df_result.columns:
        if col not in ('RBD', 'ESTABLECIMIENTO'):
            totals[col] = df_result[col].sum()
    df_result = pd.concat([df_result, pd.DataFrame([totals])], ignore_index=True)

    return df_result


def render_interactive_table(df, title="", key="itbl"):
    """Renderiza tabla HTML interactiva con selección de celdas y badge flotante de suma.

    Click = seleccionar celda. Ctrl/Cmd+Click = agregar a selección.
    Muestra badge flotante con la suma. Click afuera = deseleccionar.
    """
    meses_presentes = []
    for col in df.columns:
        if col not in ('RBD', 'ESTABLECIMIENTO'):
            parts = col.rsplit('_', 1)
            if len(parts) == 2:
                mes = parts[0]
                if mes not in meses_presentes:
                    meses_presentes.append(mes)

    # Construir header HTML multi-nivel
    header1 = '<tr><th rowspan="2" style="min-width:80px">RBD</th><th rowspan="2" style="min-width:200px">Establecimiento</th>'
    for mes in meses_presentes:
        header1 += f'<th colspan="3" class="month-header">{mes.capitalize()}</th>'
    header1 += '</tr>'

    header2 = '<tr>'
    for _ in meses_presentes:
        header2 += '<th title="Transferencia directa reconocimiento">Recon.</th>'
        header2 += '<th title="Transferencia directa tramo">Tramo</th>'
        header2 += '<th title="Asignación directa alumnos prioritarios">Prior.</th>'
    header2 += '</tr>'

    # Construir body
    body = ''
    for _, row in df.iterrows():
        is_total = str(row['RBD']).upper().startswith('TOTAL')
        tr_class = ' class="total-row"' if is_total else ''
        body += f'<tr{tr_class}>'
        body += f'<td class="rbd-cell">{html_module.escape(str(row["RBD"]))}</td>'
        body += f'<td class="name-cell">{html_module.escape(str(row["ESTABLECIMIENTO"]))}</td>'
        for mes in meses_presentes:
            for suffix in ['RECON', 'TRAMO', 'PRIOR']:
                col_name = f'{mes}_{suffix}'
                val = row.get(col_name, 0)
                val_int = int(val) if pd.notna(val) else 0
                display = fmt_clp(val_int) if val_int != 0 else '-'
                body += f'<td data-value="{val_int}">{display}</td>'
        body += '</tr>'

    n_rows = len(df)
    table_height = min(700, 90 + n_rows * 32)

    html_content = f"""
    <style>
        #{key}_wrap * {{ box-sizing: border-box; }}
        #{key}_wrap {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; font-size: 13px; }}
        .itbl {{ width: 100%; border-collapse: collapse; }}
        .itbl th {{
            background: #1e293b; color: #f1f5f9; padding: 7px 10px;
            text-align: center; font-weight: 600; font-size: 12px;
            border: 1px solid #334155; position: sticky; top: 0; z-index: 2;
        }}
        .itbl th.month-header {{ background: #334155; border-bottom: 2px solid #60a5fa; letter-spacing: 0.5px; }}
        .itbl td {{
            padding: 5px 10px; border: 1px solid #e2e8f0; cursor: pointer;
            user-select: none; text-align: right; transition: background 0.1s;
        }}
        .itbl td.rbd-cell, .itbl td.name-cell {{
            text-align: left; font-weight: 500; cursor: default; background: #f8fafc;
        }}
        .itbl tr:hover td:not(.rbd-cell):not(.name-cell) {{ background: #f0f9ff; }}
        .itbl td.selected {{ background: #dbeafe !important; outline: 2px solid #3b82f6; outline-offset: -2px; }}
        .itbl tr.total-row td {{ font-weight: 700; background: #f1f5f9 !important; border-top: 2px solid #94a3b8; }}
        .itbl tr.total-row td.selected {{ background: #bfdbfe !important; }}
        @media (prefers-color-scheme: dark) {{
            body {{ background: #0f172a; color: #e2e8f0; }}
            #{key}_wrap > div {{ border-color: #334155 !important; }}
            .itbl td {{ border-color: #334155; color: #e2e8f0; background: #0f172a; }}
            .itbl td.rbd-cell, .itbl td.name-cell {{ background: #1e293b; color: #e2e8f0; }}
            .itbl tr:hover td:not(.rbd-cell):not(.name-cell) {{ background: #172554; }}
            .itbl td.selected {{ background: #1e3a5f !important; outline-color: #60a5fa; }}
            .itbl tr.total-row td {{ background: #1e293b !important; border-top-color: #475569; color: #f1f5f9; }}
            .itbl tr.total-row td.selected {{ background: #1e40af !important; }}
        }}
        #sumBadge_{key} {{
            position: fixed; bottom: 24px; right: 24px;
            background: linear-gradient(135deg, #1e293b, #334155); color: white;
            padding: 14px 22px; border-radius: 14px;
            font-size: 15px; font-weight: 600;
            z-index: 99999; box-shadow: 0 8px 30px rgba(0,0,0,0.3);
            display: none; backdrop-filter: blur(8px);
        }}
        #sumBadge_{key} .count {{ opacity: 0.6; font-size: 12px; margin-right: 6px; }}
        @keyframes fadeUp {{ from {{ opacity:0; transform:translateY(10px); }} to {{ opacity:1; transform:translateY(0); }} }}
        #sumBadge_{key}.visible {{ display: block; animation: fadeUp 0.2s ease; }}
    </style>
    <div id="{key}_wrap">
        <div style="overflow:auto; max-height:{table_height}px; border-radius:8px; border:1px solid #e2e8f0;">
            <table class="itbl" id="tbl_{key}">
                <thead>{header1}{header2}</thead>
                <tbody>{body}</tbody>
            </table>
        </div>
    </div>
    <div id="sumBadge_{key}"></div>
    <script>
    (function() {{
        const tbl = document.getElementById('tbl_{key}');
        const badge = document.getElementById('sumBadge_{key}');
        const cells = tbl.querySelectorAll('td[data-value]');
        let sel = new Set();

        function fmtCLP(n) {{
            let s = Math.round(Math.abs(n)).toString();
            let r = '';
            for (let i = s.length - 1, c = 0; i >= 0; i--, c++) {{
                if (c > 0 && c % 3 === 0) r = '.' + r;
                r = s[i] + r;
            }}
            return (n < 0 ? '-' : '') + '$' + r;
        }}

        function refresh() {{
            if (sel.size === 0) {{
                badge.classList.remove('visible');
                badge.style.display = 'none';
                return;
            }}
            let sum = 0;
            sel.forEach(c => {{ sum += parseFloat(c.dataset.value) || 0; }});
            badge.innerHTML = '<span class="count">&Sigma; (' + sel.size + ')</span> ' + fmtCLP(sum);
            badge.style.display = 'block';
            badge.classList.add('visible');
        }}

        cells.forEach(cell => {{
            cell.addEventListener('click', e => {{
                e.stopPropagation();
                if (e.ctrlKey || e.metaKey) {{
                    if (cell.classList.contains('selected')) {{
                        cell.classList.remove('selected');
                        sel.delete(cell);
                    }} else {{
                        cell.classList.add('selected');
                        sel.add(cell);
                    }}
                }} else {{
                    sel.forEach(c => c.classList.remove('selected'));
                    sel.clear();
                    cell.classList.add('selected');
                    sel.add(cell);
                }}
                refresh();
            }});
        }});

        document.addEventListener('click', e => {{
            if (!e.target.closest('#tbl_{key}') && !e.target.closest('#sumBadge_{key}')) {{
                sel.forEach(c => c.classList.remove('selected'));
                sel.clear();
                refresh();
            }}
        }});
    }})();
    </script>
    """
    st_components.html(html_content, height=table_height + 50, scrolling=False)


def render_generic_interactive_table(df, money_cols=None, label_cols=None, key="gtbl"):
    """Renderiza tabla HTML interactiva genérica con selección de celdas y badge flotante de suma.

    Args:
        df: DataFrame a renderizar.
        money_cols: set/list de columnas monetarias (si None, se auto-detecta).
        label_cols: set/list de columnas de texto (no seleccionables).
        key: clave única para el componente.
    """
    if money_cols is None:
        money_cols = set()
        for col in df.columns:
            if col not in NON_MONEY_COLS and df[col].dtype in ('float64', 'int64', 'float32', 'int32'):
                money_cols.add(col)
    else:
        money_cols = set(money_cols)

    if label_cols is None:
        label_cols = set()
        for col in df.columns:
            if df[col].dtype == 'object' or col in NON_MONEY_COLS:
                label_cols.add(col)
    else:
        label_cols = set(label_cols)

    # Header
    header = '<tr>'
    for col in df.columns:
        header += f'<th>{html_module.escape(str(col))}</th>'
    header += '</tr>'

    # Body
    body = ''
    for _, row in df.iterrows():
        is_total = False
        for col in df.columns:
            val_str = str(row[col]).upper().strip()
            if val_str.startswith('TOTAL'):
                is_total = True
                break
        tr_class = ' class="total-row"' if is_total else ''
        body += f'<tr{tr_class}>'
        for col in df.columns:
            val = row[col]
            if col in label_cols:
                display = html_module.escape(str(val)) if pd.notna(val) else ''
                body += f'<td class="label-cell">{display}</td>'
            elif col in money_cols:
                val_int = int(val) if pd.notna(val) else 0
                display = fmt_clp(val_int) if val_int != 0 else '-'
                body += f'<td data-value="{val_int}">{display}</td>'
            else:
                # Numérica no-monetaria (horas, conteos)
                if pd.notna(val):
                    try:
                        display = str(int(val)) if float(val) == int(float(val)) else str(val)
                    except (ValueError, TypeError):
                        display = html_module.escape(str(val))
                else:
                    display = ''
                body += f'<td class="label-cell">{display}</td>'
        body += '</tr>'

    n_rows = len(df)
    table_height = min(700, 90 + n_rows * 32)

    html_content = f"""
    <style>
        #{key}_wrap * {{ box-sizing: border-box; }}
        #{key}_wrap {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; font-size: 13px; }}
        #{key}_wrap .gtbl {{ width: 100%; border-collapse: collapse; }}
        #{key}_wrap .gtbl th {{
            background: #1e293b; color: #f1f5f9; padding: 7px 10px;
            text-align: center; font-weight: 600; font-size: 12px;
            border: 1px solid #334155; position: sticky; top: 0; z-index: 2;
        }}
        #{key}_wrap .gtbl td {{
            padding: 5px 10px; border: 1px solid #e2e8f0; cursor: pointer;
            user-select: none; text-align: right; transition: background 0.1s;
        }}
        #{key}_wrap .gtbl td.label-cell {{
            text-align: left; font-weight: 500; cursor: default; background: #f8fafc;
        }}
        #{key}_wrap .gtbl tr:hover td:not(.label-cell) {{ background: #f0f9ff; }}
        #{key}_wrap .gtbl td.selected {{ background: #dbeafe !important; outline: 2px solid #3b82f6; outline-offset: -2px; }}
        #{key}_wrap .gtbl tr.total-row td {{ font-weight: 700; background: #f1f5f9 !important; border-top: 2px solid #94a3b8; }}
        #{key}_wrap .gtbl tr.total-row td.selected {{ background: #bfdbfe !important; }}
        @media (prefers-color-scheme: dark) {{
            body {{ background: #0f172a; color: #e2e8f0; }}
            #{key}_wrap > div {{ border-color: #334155 !important; }}
            #{key}_wrap .gtbl td {{ border-color: #334155; color: #e2e8f0; background: #0f172a; }}
            #{key}_wrap .gtbl td.label-cell {{ background: #1e293b; color: #e2e8f0; }}
            #{key}_wrap .gtbl tr:hover td:not(.label-cell) {{ background: #172554; }}
            #{key}_wrap .gtbl td.selected {{ background: #1e3a5f !important; outline-color: #60a5fa; }}
            #{key}_wrap .gtbl tr.total-row td {{ background: #1e293b !important; border-top-color: #475569; color: #f1f5f9; }}
            #{key}_wrap .gtbl tr.total-row td.selected {{ background: #1e40af !important; }}
        }}
        #sumBadge_{key} {{
            position: fixed; bottom: 24px; right: 24px;
            background: linear-gradient(135deg, #1e293b, #334155); color: white;
            padding: 14px 22px; border-radius: 14px;
            font-size: 15px; font-weight: 600;
            z-index: 99999; box-shadow: 0 8px 30px rgba(0,0,0,0.3);
            display: none; backdrop-filter: blur(8px);
        }}
        #sumBadge_{key} .count {{ opacity: 0.6; font-size: 12px; margin-right: 6px; }}
        @keyframes fadeUp_{key} {{ from {{ opacity:0; transform:translateY(10px); }} to {{ opacity:1; transform:translateY(0); }} }}
        #sumBadge_{key}.visible {{ display: block; animation: fadeUp_{key} 0.2s ease; }}
    </style>
    <div id="{key}_wrap">
        <div style="overflow:auto; max-height:{table_height}px; border-radius:8px; border:1px solid #e2e8f0;">
            <table class="gtbl" id="tbl_{key}">
                <thead>{header}</thead>
                <tbody>{body}</tbody>
            </table>
        </div>
    </div>
    <div id="sumBadge_{key}"></div>
    <script>
    (function() {{
        const tbl = document.getElementById('tbl_{key}');
        const badge = document.getElementById('sumBadge_{key}');
        const cells = tbl.querySelectorAll('td[data-value]');
        let sel = new Set();

        function fmtCLP(n) {{
            let s = Math.round(Math.abs(n)).toString();
            let r = '';
            for (let i = s.length - 1, c = 0; i >= 0; i--, c++) {{
                if (c > 0 && c % 3 === 0) r = '.' + r;
                r = s[i] + r;
            }}
            return (n < 0 ? '-' : '') + '$' + r;
        }}

        function refresh() {{
            if (sel.size === 0) {{
                badge.classList.remove('visible');
                badge.style.display = 'none';
                return;
            }}
            let sum = 0;
            sel.forEach(c => {{ sum += parseFloat(c.dataset.value) || 0; }});
            badge.innerHTML = '<span class="count">&Sigma; (' + sel.size + ')</span> ' + fmtCLP(sum);
            badge.style.display = 'block';
            badge.classList.add('visible');
        }}

        cells.forEach(cell => {{
            cell.addEventListener('click', e => {{
                e.stopPropagation();
                if (e.ctrlKey || e.metaKey) {{
                    if (cell.classList.contains('selected')) {{
                        cell.classList.remove('selected');
                        sel.delete(cell);
                    }} else {{
                        cell.classList.add('selected');
                        sel.add(cell);
                    }}
                }} else {{
                    sel.forEach(c => c.classList.remove('selected'));
                    sel.clear();
                    cell.classList.add('selected');
                    sel.add(cell);
                }}
                refresh();
            }});
        }});

        document.addEventListener('click', e => {{
            if (!e.target.closest('#tbl_{key}') && !e.target.closest('#sumBadge_{key}')) {{
                sel.forEach(c => c.classList.remove('selected'));
                sel.clear();
                refresh();
            }}
        }});
    }})();
    </script>
    """
    st_components.html(html_content, height=table_height + 50, scrolling=False)


st.set_page_config(
    page_title="RemuPro",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# ESTILOS CSS ADAPTATIVOS (Claro/Oscuro)
# ============================================================================

st.markdown("""
<style>
    /* ===== RESET Y BASE ===== */
    .block-container {
        padding: 2rem 3rem !important;
        max-width: 1200px;
    }
    
    /* ===== HEADER ===== */
    .app-header {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding-bottom: 1.5rem;
        margin-bottom: 1.5rem;
        border-bottom: 1px solid var(--text-color);
        opacity: 0.9;
    }
    
    .app-title {
        font-size: 2rem;
        font-weight: 700;
        margin: 0;
    }
    
    .app-subtitle {
        font-size: 0.95rem;
        opacity: 0.7;
        margin: 0.25rem 0 0 0;
    }
    
    .app-version {
        font-size: 0.85rem;
        opacity: 0.5;
    }
    
    /* ===== CARDS ===== */
    .custom-card {
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1.25rem;
        border: 1px solid rgba(128, 128, 128, 0.2);
        background: rgba(128, 128, 128, 0.05);
    }
    
    .card-header {
        font-size: 1.1rem;
        font-weight: 600;
        margin-bottom: 1rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    
    .card-content {
        font-size: 0.95rem;
        line-height: 1.6;
        opacity: 0.85;
    }
    
    /* ===== INFO BOXES ===== */
    .info-box {
        background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
        color: white;
        padding: 1rem 1.25rem;
        border-radius: 10px;
        margin-bottom: 1.25rem;
        font-size: 0.9rem;
    }
    
    .success-box {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        padding: 1rem 1.25rem;
        border-radius: 10px;
        margin: 1rem 0;
        font-size: 0.95rem;
    }
    
    .warning-box {
        background: rgba(245, 158, 11, 0.15);
        border: 1px solid #f59e0b;
        color: #f59e0b;
        padding: 1rem 1.25rem;
        border-radius: 10px;
        margin: 1rem 0;
        font-size: 0.9rem;
    }
    
    /* ===== TABS ===== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.5rem;
        background: rgba(128, 128, 128, 0.1);
        padding: 0.5rem;
        border-radius: 12px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        padding: 0.75rem 1.5rem;
        font-weight: 500;
    }
    
    .stTabs [aria-selected="true"] {
        background: #3b82f6 !important;
        color: white !important;
    }
    
    /* ===== BOTONES ===== */
    .stButton > button {
        width: 100%;
        padding: 0.875rem 1.5rem;
        font-weight: 600;
        font-size: 1rem;
        border-radius: 10px;
        border: none;
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        transition: all 0.2s ease;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(16, 185, 129, 0.3);
    }
    
    .stButton > button:active {
        transform: translateY(0);
    }
    
    .stDownloadButton > button {
        background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%) !important;
        color: white !important;
        border: none !important;
        width: 100%;
        padding: 0.875rem 1.5rem;
        font-weight: 600;
        border-radius: 10px;
    }
    
    .stDownloadButton > button:hover {
        box-shadow: 0 8px 20px rgba(59, 130, 246, 0.3);
    }
    
    /* ===== FILE UPLOADER ===== */
    [data-testid="stFileUploader"] {
        border-radius: 10px;
    }
    
    [data-testid="stFileUploader"] section {
        border: 2px dashed rgba(128, 128, 128, 0.3);
        border-radius: 10px;
        padding: 1.5rem;
        transition: all 0.2s ease;
    }
    
    [data-testid="stFileUploader"] section:hover {
        border-color: #3b82f6;
        background: rgba(59, 130, 246, 0.05);
    }
    
    /* ===== METRICS ===== */
    [data-testid="stMetricValue"] {
        font-size: 1.75rem !important;
        font-weight: 700 !important;
        color: #3b82f6 !important;
    }
    
    [data-testid="stMetricLabel"] {
        font-weight: 500;
    }
    
    /* ===== EXPANDER ===== */
    .streamlit-expanderHeader {
        font-weight: 600;
        font-size: 1rem;
        border-radius: 10px;
    }
    
    /* ===== SELECTBOX ===== */
    .stSelectbox > div > div {
        border-radius: 8px;
    }
    
    /* ===== PROGRESS ===== */
    .stProgress > div > div {
        background: linear-gradient(90deg, #3b82f6 0%, #10b981 100%);
        border-radius: 10px;
    }
    
    /* ===== DATAFRAME ===== */
    [data-testid="stDataFrame"] {
        border-radius: 10px;
        overflow: hidden;
    }
    
    /* ===== FOOTER ===== */
    .app-footer {
        text-align: center;
        padding: 2rem 0 1rem 0;
        margin-top: 2rem;
        border-top: 1px solid rgba(128, 128, 128, 0.2);
        font-size: 0.85rem;
        opacity: 0.6;
    }
    
    /* ===== AUDIT LOG ===== */
    .audit-summary-bar {
        display: flex;
        gap: 1rem;
        padding: 0.75rem 1rem;
        border-radius: 10px;
        background: rgba(128, 128, 128, 0.08);
        border: 1px solid rgba(128, 128, 128, 0.15);
        margin-bottom: 1rem;
        flex-wrap: wrap;
        align-items: center;
    }
    .audit-summary-bar .audit-stat {
        display: flex;
        align-items: center;
        gap: 0.4rem;
        font-size: 0.9rem;
        font-weight: 500;
    }
    .audit-badge {
        display: inline-block;
        padding: 0.15rem 0.6rem;
        border-radius: 999px;
        font-size: 0.75rem;
        font-weight: 600;
        color: white;
        line-height: 1.4;
    }
    .audit-badge-error { background: #ef4444; }
    .audit-badge-warning { background: #f59e0b; }
    .audit-badge-info { background: #3b82f6; }
    .audit-badge-tipo {
        background: rgba(128, 128, 128, 0.15);
        color: inherit;
        font-weight: 500;
    }
    .audit-group-header {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        padding: 0.5rem 0;
        font-weight: 600;
        font-size: 0.95rem;
        border-bottom: 1px solid rgba(128, 128, 128, 0.15);
        margin-bottom: 0.5rem;
    }
    .audit-entry {
        display: flex;
        gap: 0.75rem;
        padding: 0.4rem 0.5rem;
        border-radius: 6px;
        font-size: 0.85rem;
        align-items: flex-start;
    }
    .audit-entry:hover {
        background: rgba(128, 128, 128, 0.06);
    }
    .audit-entry .audit-time {
        opacity: 0.5;
        min-width: 65px;
        font-family: monospace;
        font-size: 0.8rem;
    }
    .audit-entry .audit-msg {
        flex: 1;
        line-height: 1.4;
    }
    .audit-duration {
        font-size: 0.85rem;
        opacity: 0.6;
        margin-top: 0.5rem;
        font-style: italic;
    }

    /* ===== OCULTAR ELEMENTOS DE STREAMLIT ===== */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* ===== ANIMACIONES ===== */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(10px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .custom-card {
        animation: fadeIn 0.3s ease-out;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def show_header():
    """Muestra el header de la aplicación."""
    st.markdown(f"""
    <div class="app-header">
        <div>
            <h1 class="app-title">📊 RemuPro</h1>
            <p class="app-subtitle">Sistema de Procesamiento de Remuneraciones Educativas</p>
        </div>
        <div class="app-version">v{VERSION}</div>
    </div>
    """, unsafe_allow_html=True)


def info_box(text: str):
    """Muestra una caja de información (HTML-sanitized)."""
    safe_text = _sanitize_html(text)
    st.markdown(f'<div class="info-box">{safe_text}</div>', unsafe_allow_html=True)


def success_box(text: str):
    """Muestra una caja de éxito (HTML-sanitized)."""
    safe_text = _sanitize_html(text)
    st.markdown(f'<div class="success-box">{safe_text}</div>', unsafe_allow_html=True)


def warning_box(text: str):
    """Muestra una caja de advertencia (HTML-sanitized)."""
    safe_text = _sanitize_html(text)
    st.markdown(f'<div class="warning-box">{safe_text}</div>', unsafe_allow_html=True)


def show_audit_log_detailed(audit):
    """
    Muestra el log de auditoría con resumen visual, agrupado por tipo,
    ordenado por severidad (errores primero), y con info de duración.
    """
    if not audit or len(audit) == 0:
        return

    summary = audit.get_summary()
    total = summary.get('total', 0)
    n_errors = summary.get('errores', 0)
    n_warnings = summary.get('advertencias', 0)
    n_info = total - n_errors - n_warnings

    # --- Summary bar ---
    parts = [f'<span class="audit-stat">Total: <b>{total}</b></span>']
    if n_errors:
        parts.append(
            f'<span class="audit-stat"><span class="audit-badge audit-badge-error">'
            f'ERROR</span> {n_errors}</span>'
        )
    if n_warnings:
        parts.append(
            f'<span class="audit-stat"><span class="audit-badge audit-badge-warning">'
            f'WARNING</span> {n_warnings}</span>'
        )
    if n_info:
        parts.append(
            f'<span class="audit-stat"><span class="audit-badge audit-badge-info">'
            f'INFO</span> {n_info}</span>'
        )

    # Duration
    duration_text = ""
    proceso_entries = audit.get_by_tipo(audit.TIPO_PROCESO)
    if len(proceso_entries) >= 2:
        t_start = proceso_entries[0].timestamp
        t_end = proceso_entries[-1].timestamp
        secs = (t_end - t_start).total_seconds()
        duration_text = f' <span class="audit-stat" style="margin-left:auto;">Duracion: <b>{secs:.1f}s</b></span>'

    st.markdown(
        f'<div class="audit-summary-bar">{"".join(parts)}{duration_text}</div>',
        unsafe_allow_html=True
    )

    # --- Build sorted entries: errors > warnings > info ---
    nivel_order = {'ERROR': 0, 'WARNING': 1, 'INFO': 2}
    sorted_entries = sorted(
        audit.entries,
        key=lambda e: (nivel_order.get(e.nivel, 3), e.timestamp)
    )

    # --- Friendly type names ---
    tipo_labels = {
        'columna_faltante': 'Columnas Faltantes',
        'valor_inusual': 'Valores Inusuales',
        'docente_eib': 'Docentes EIB',
        'excede_horas': 'Excede Horas',
        'sin_liquidacion': 'Sin Liquidacion',
        'validacion': 'Validacion',
        'proceso': 'Proceso',
        'archivo': 'Archivo',
    }

    badge_class = {
        'ERROR': 'audit-badge-error',
        'WARNING': 'audit-badge-warning',
        'INFO': 'audit-badge-info',
    }

    # --- Expander with grouped entries ---
    label = "Ver log de auditoria detallado"
    if n_errors:
        label = f"Log de auditoria ({n_errors} errores, {n_warnings} advertencias)"
    elif n_warnings:
        label = f"Log de auditoria ({n_warnings} advertencias)"

    with st.expander(label):
        # Filter
        niveles_disponibles = sorted(
            set(e.nivel for e in audit.entries),
            key=lambda n: nivel_order.get(n, 3)
        )
        filtro = st.multiselect(
            "Filtrar por nivel",
            options=niveles_disponibles,
            default=niveles_disponibles,
            key="audit_log_filter"
        )

        filtered = [e for e in sorted_entries if e.nivel in filtro]

        if not filtered:
            st.info("No hay entradas con los filtros seleccionados.")
            return

        # Group by tipo
        from collections import OrderedDict
        groups = OrderedDict()
        for entry in filtered:
            tipo_key = entry.tipo
            if tipo_key not in groups:
                groups[tipo_key] = []
            groups[tipo_key].append(entry)

        for tipo_key, entries in groups.items():
            tipo_label = tipo_labels.get(tipo_key, tipo_key.replace('_', ' ').title())
            # Determine worst level in group for header badge
            worst = 'INFO'
            for e in entries:
                if e.nivel == 'ERROR':
                    worst = 'ERROR'
                    break
                if e.nivel == 'WARNING':
                    worst = 'WARNING'

            badge_cls = badge_class.get(worst, 'audit-badge-info')
            header_html = (
                f'<div class="audit-group-header">'
                f'<span class="audit-badge {badge_cls}">{worst}</span> '
                f'{tipo_label} ({len(entries)})'
                f'</div>'
            )
            st.markdown(header_html, unsafe_allow_html=True)

            # Render entries (limit to 50 per group)
            display_entries = entries[:50]
            rows_html = []
            for e in display_entries:
                time_str = e.timestamp.strftime("%H:%M:%S")
                bcls = badge_class.get(e.nivel, 'audit-badge-info')
                rows_html.append(
                    f'<div class="audit-entry">'
                    f'<span class="audit-time">{time_str}</span>'
                    f'<span class="audit-badge {bcls}">{e.nivel}</span>'
                    f'<span class="audit-msg">{_sanitize_html(e.mensaje)}</span>'
                    f'</div>'
                )
            st.markdown("".join(rows_html), unsafe_allow_html=True)

            if len(entries) > 50:
                st.caption(f"... y {len(entries) - 50} entradas mas en este grupo")

        # Table view
        st.markdown("---")
        st.markdown("**Tabla completa**")
        df_audit = pd.DataFrame([
            {
                'Hora': e.timestamp.strftime("%H:%M:%S"),
                'Nivel': e.nivel,
                'Tipo': tipo_labels.get(e.tipo, e.tipo),
                'Mensaje': e.mensaje,
            }
            for e in filtered
        ])
        st.dataframe(df_audit, width='stretch', hide_index=True)


def format_user_error(e: Exception) -> str:
    """Traduce excepciones técnicas a mensajes amigables en español."""
    error_str = str(e)
    error_type = type(e).__name__

    if error_type == 'ColumnMissingError' or 'columna' in error_str.lower():
        return "El archivo no tiene las columnas esperadas. Verifique que usa el archivo correcto."
    elif error_type == 'FileValidationError' or 'validación' in error_str.lower():
        if 'vacío' in error_str.lower() or 'empty' in error_str.lower():
            return "El archivo está vacío. Verifique que el archivo tiene datos."
        elif 'formato' in error_str.lower() or 'format' in error_str.lower():
            return "El formato del archivo no es válido. Use un archivo Excel (.xlsx) o CSV (.csv)."
        elif 'no encontr' in error_str.lower() or 'not found' in error_str.lower():
            return "No se encontró el archivo. Vuelva a cargarlo."
        return f"Error de validación: {error_str}"
    elif isinstance(e, PermissionError) or 'permission' in error_str.lower():
        return "El archivo está abierto en otro programa. Cierre el archivo en Excel e intente nuevamente."
    else:
        return f"Ocurrió un error. Detalle: {error_str}"


def add_school_names(df, rbd_col='RBD'):
    """Agrega columna ESCUELA al DataFrame basándose en el RBD."""
    escuelas = get_rbd_map()
    if not escuelas or rbd_col not in df.columns:
        return df
    df = df.copy()
    df['ESCUELA'] = df[rbd_col].astype(str).str.replace(r'\.0$', '', regex=True).map(
        lambda x: escuelas.get(str(x).split('.')[0].split('-')[0].strip(), '')
    )
    # Mover ESCUELA después de RBD
    cols = list(df.columns)
    if 'ESCUELA' in cols and rbd_col in cols:
        cols.remove('ESCUELA')
        idx = cols.index(rbd_col) + 1
        cols.insert(idx, 'ESCUELA')
        df = df[cols]
    return df


def show_column_alerts(column_alerts):
    """Muestra alertas de columnas con lista expandible para columnas nuevas."""
    if not column_alerts:
        return
    criticas = [a for a in column_alerts if a['nivel'] == 'error']
    nuevas = [a for a in column_alerts if a['tipo'] == 'columna_nueva']

    if criticas:
        for a in criticas:
            warning_box(f"<b>Columna no encontrada:</b> {a['columna_nombre']}<br>"
                        f"Los montos de este concepto serán $0.")
    if nuevas:
        for a in nuevas:
            lista = a.get('columnas_lista', [])
            if lista:
                info_box(f"<b>Columnas nuevas detectadas ({len(lista)}):</b> "
                         f"El archivo MINEDUC tiene columnas que el sistema no procesa.")
                with st.expander(f"Ver las {len(lista)} columnas nuevas"):
                    for col in lista:
                        st.markdown(f"- `{col}`")
                    st.caption("Para registrar estas columnas, agregarlas en "
                               "`config/columns.py` → `WEB_SOSTENEDOR_COLUMNS`")


def show_desglose_daem_cpeip(df, prefix=''):
    """Muestra desglose DAEM/CPEIP con totales detallados por concepto."""
    cols = df.columns

    daem_recon = sum(df[f'DAEM_RECON_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL'] if f'DAEM_RECON_{s}' in cols)
    daem_tramo = sum(df[f'DAEM_TRAMO_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL'] if f'DAEM_TRAMO_{s}' in cols)
    daem_total = daem_recon + daem_tramo

    cpeip_recon = sum(df[f'CPEIP_RECON_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL'] if f'CPEIP_RECON_{s}' in cols)
    cpeip_tramo = sum(df[f'CPEIP_TRAMO_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL'] if f'CPEIP_TRAMO_{s}' in cols)
    cpeip_prior = sum(df[f'CPEIP_PRIOR_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL'] if f'CPEIP_PRIOR_{s}' in cols)
    cpeip_total = cpeip_recon + cpeip_tramo + cpeip_prior

    st.markdown("##### 🏛️ Desglose por Pagador")
    st.caption("DAEM = lo que paga el municipio (Subvención) | CPEIP = lo que transfiere el ministerio (Transferencia)")

    col_daem, col_cpeip = st.columns(2)
    with col_daem:
        st.metric("Total DAEM (Subvención)", fmt_clp(daem_total))
        c1, c2 = st.columns(2)
        c1.metric("DAEM Reconocimiento", fmt_clp(daem_recon))
        c2.metric("DAEM Tramo", fmt_clp(daem_tramo))
    with col_cpeip:
        st.metric("Total CPEIP (Transferencia)", fmt_clp(cpeip_total))
        c1, c2, c3 = st.columns(3)
        c1.metric("CPEIP Reconocimiento", fmt_clp(cpeip_recon))
        c2.metric("CPEIP Tramo", fmt_clp(cpeip_tramo))
        c3.metric("CPEIP Prioritarios", fmt_clp(cpeip_prior))


def show_revision_table(df_revision):
    """Muestra tabla de revisión con RUT formateado y nombre/apellido visibles."""
    if df_revision.empty:
        return

    # Formatear RUT con guión
    if 'RUT' in df_revision.columns:
        df_revision = df_revision.copy()
        df_revision['RUT'] = df_revision['RUT'].apply(format_rut)

    # Asegurar que NOMBRE y APELLIDOS estén visibles al inicio
    cols_prioritarias = ['RUT', 'NOMBRE', 'APELLIDOS', 'TIPO_PAGO', 'MOTIVO',
                         'HORAS_SEP', 'HORAS_PIE', 'HORAS_SN', 'HORAS_TOTAL',
                         'EXCESO', 'DETALLE', 'ACCION']
    cols_exist = [c for c in cols_prioritarias if c in df_revision.columns]
    cols_resto = [c for c in df_revision.columns if c not in cols_exist]
    df_revision = df_revision[cols_exist + cols_resto]

    st.dataframe(df_revision, width='stretch', hide_index=True)


def show_charts_by_school(df_rbd):
    """Muestra gráficos de distribución BRP por escuela."""
    import plotly.express as px
    import plotly.graph_objects as go

    # Agregar nombres de escuela
    df_rbd = add_school_names(df_rbd)

    # Excluir fila TOTAL
    df_chart = df_rbd[df_rbd['RBD'].astype(str) != 'TOTAL'].copy()
    if df_chart.empty:
        return

    # Etiqueta para eje: ESCUELA o RBD
    if 'ESCUELA' in df_chart.columns:
        df_chart['LABEL'] = df_chart.apply(
            lambda r: r['ESCUELA'] if r['ESCUELA'] else str(r['RBD']), axis=1
        )
    else:
        df_chart['LABEL'] = df_chart['RBD'].astype(str)

    # Gráfico de barras apiladas: SEP/PIE/NORMAL por escuela
    fig = go.Figure()
    for subv, color in [('BRP_SEP', '#3b82f6'), ('BRP_PIE', '#10b981'), ('BRP_NORMAL', '#f59e0b')]:
        if subv in df_chart.columns:
            fig.add_trace(go.Bar(
                name=subv.replace('BRP_', ''),
                x=df_chart['LABEL'],
                y=df_chart[subv],
                marker_color=color,
                text=df_chart[subv].apply(lambda x: fmt_clp(x)),
                textposition='inside'
            ))
    fig.update_layout(
        barmode='stack',
        title='BRP Total por Escuela (SEP / PIE / Normal)',
        xaxis_title='',
        yaxis_title='Monto ($)',
        height=400,
        xaxis_tickangle=-35,
        showlegend=True,
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        separators=',.'
    )
    st.plotly_chart(fig, width='stretch')

    # Gráfico DAEM vs CPEIP por escuela
    daem_cols = ['DAEM_SEP', 'DAEM_PIE', 'DAEM_NORMAL']
    cpeip_cols = ['CPEIP_SEP', 'CPEIP_PIE', 'CPEIP_NORMAL']
    has_daem = all(c in df_chart.columns for c in daem_cols)
    has_cpeip = all(c in df_chart.columns for c in cpeip_cols)

    if has_daem and has_cpeip:
        df_chart['TOTAL_DAEM'] = df_chart[daem_cols].sum(axis=1)
        df_chart['TOTAL_CPEIP'] = df_chart[cpeip_cols].sum(axis=1)

        fig2 = go.Figure()
        fig2.add_trace(go.Bar(
            name='DAEM', x=df_chart['LABEL'], y=df_chart['TOTAL_DAEM'],
            marker_color='#6366f1',
            text=df_chart['TOTAL_DAEM'].apply(lambda x: fmt_clp(x)),
            textposition='inside'
        ))
        fig2.add_trace(go.Bar(
            name='CPEIP', x=df_chart['LABEL'], y=df_chart['TOTAL_CPEIP'],
            marker_color='#ec4899',
            text=df_chart['TOTAL_CPEIP'].apply(lambda x: fmt_clp(x)),
            textposition='inside'
        ))
        fig2.update_layout(
            barmode='group',
            title='DAEM vs CPEIP por Escuela',
            xaxis_title='',
            yaxis_title='Monto ($)',
            height=400,
            xaxis_tickangle=-35,
            showlegend=True,
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
            separators=',.'
        )
        st.plotly_chart(fig2, width='stretch')


def show_rem_results(df_resumen, alertas):
    """Muestra resultados del procesamiento REM (horas disponibles)."""
    st.markdown("##### 📋 Análisis de Horas (REM)")

    # Métricas generales
    total_personas = len(df_resumen)
    exceden = df_resumen['EXCEDE'].sum() if 'EXCEDE' in df_resumen.columns else 0
    total_horas = df_resumen['TOTAL'].sum() if 'TOTAL' in df_resumen.columns else 0
    disponibles = df_resumen['DISPONIBLE'].sum() if 'DISPONIBLE' in df_resumen.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Personas", f"{total_personas}")
    with c2:
        st.metric("Total Horas", f"{int(total_horas)}")
    with c3:
        st.metric("Horas Disponibles", f"{int(disponibles)}")
    with c4:
        st.metric("Exceden 44 hrs", f"{int(exceden)}")
        if exceden > 0:
            st.badge(f"{exceden} alerta(s)", color="red")

    # Alertas de exceso
    if alertas:
        st.markdown("---")
        st.markdown(f"##### ⚠️ {len(alertas)} persona(s) exceden 44 horas")
        for a in alertas:
            warning_box(
                f"<b>{a.get('nombre', a['rut'])}</b> — {a['detalle']}<br>"
                f"Exceso: <b>{a['exceso']} hrs</b>"
            )

    # Tabla resumen con RUT formateado
    st.markdown("---")
    st.markdown("##### 📊 Detalle por persona")

    df_display = df_resumen.copy()
    if 'RUT_NORM' in df_display.columns:
        df_display['RUT'] = df_display['RUT_NORM'].apply(format_rut)
        cols = ['RUT'] + [c for c in df_display.columns if c not in ('RUT', 'RUT_NORM')]
        df_display = df_display[cols]
        df_display = df_display.drop(columns=['RUT_NORM'], errors='ignore')

    # Resaltar filas que exceden
    st.dataframe(df_display, width='stretch', hide_index=True)

    # Resumen por tipo
    if total_horas > 0:
        import plotly.express as px
        horas_sep = df_resumen['SEP'].sum() if 'SEP' in df_resumen.columns else 0
        horas_pie = df_resumen['PIE'].sum() if 'PIE' in df_resumen.columns else 0
        horas_normal = df_resumen['NORMAL'].sum() if 'NORMAL' in df_resumen.columns else 0
        horas_eib = df_resumen['EIB'].sum() if 'EIB' in df_resumen.columns else 0

        tipos = ['SEP', 'PIE', 'Normal', 'EIB']
        vals = [horas_sep, horas_pie, horas_normal, horas_eib]
        # Filtrar los que tienen valor
        data = [(t, v) for t, v in zip(tipos, vals) if v > 0]
        if data:
            df_h = pd.DataFrame(data, columns=['Tipo', 'Horas'])
            fig = px.pie(df_h, values='Horas', names='Tipo',
                         title='Distribución de Horas por Tipo',
                         color_discrete_sequence=['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6'])
            fig.update_traces(textposition='inside', textinfo='percent+label')
            fig.update_layout(height=300, showlegend=False, separators=',.')
            st.plotly_chart(fig, width='stretch')


def show_multi_establishment(excel_source):
    """Muestra sección de docentes multi-establecimiento.

    Args:
        excel_source: Path al archivo Excel, o bytes del Excel.
    """
    try:
        src = BytesIO(excel_source) if isinstance(excel_source, bytes) else excel_source
        df_multi = pd.read_excel(src, sheet_name='MULTI_ESTABLECIMIENTO', engine='openpyxl')
    except Exception:
        return

    if df_multi.empty:
        return

    total_docentes = df_multi[df_multi['TIPO_FILA'] == 'TOTAL_DOCENTE']
    n_docentes = len(total_docentes)

    st.markdown("---")
    st.markdown("##### 🏫 Docentes Multi-Establecimiento")
    st.caption(f"{n_docentes} docente(s) trabajan en 2 o más escuelas")

    if n_docentes == 0:
        st.info("No hay docentes en múltiples establecimientos.")
        return

    # Agregar nombres de escuela a los RBD
    escuelas = get_rbd_map()

    # Formatear RUT
    df_multi = df_multi.copy()
    df_multi['RUT'] = df_multi['RUT'].apply(format_rut)

    # Agregar nombre escuela
    if escuelas:
        df_multi['ESCUELA'] = df_multi['RBD'].astype(str).map(
            lambda x: escuelas.get(str(x).split('.')[0].split('-')[0].strip(), '') if x != 'TOTAL' else ''
        )

    # Tabs: Resumen general + cada docente
    tab_labels = ["📋 Resumen"]
    for _, total_row in total_docentes.iterrows():
        nombre = total_row.get('NOMBRE', '')
        short = nombre.split()[0] if nombre and ' ' in str(nombre) else str(nombre)[:15]
        tab_labels.append(f"👤 {short}")

    multi_tabs = st.tabs(tab_labels)

    # Tab 0: Resumen general con buscador
    with multi_tabs[0]:
        buscar_multi = st.text_input("🔎 Buscar docente", key="multi_buscar")
        df_totals_display = total_docentes.copy()
        cols_total = ['RUT', 'NOMBRE', 'HORAS_CONTRATO', 'RECONOCIMIENTO_MINEDUC',
                      'TRAMO_MINEDUC', 'PRIORITARIOS_MINEDUC', 'BRP_TOTAL']
        cols_total = [c for c in cols_total if c in df_totals_display.columns]

        if buscar_multi.strip():
            term = buscar_multi.strip().lower()
            mask = (
                df_totals_display['RUT'].astype(str).str.lower().str.contains(term, na=False) |
                df_totals_display['NOMBRE'].astype(str).str.lower().str.contains(term, na=False)
            )
            df_totals_display = df_totals_display[mask]

        st.dataframe(df_totals_display[cols_total], width='stretch', hide_index=True)
        st.caption(f"Mostrando {len(df_totals_display)} de {n_docentes} docentes")

    # Tab por cada docente
    for tab_idx, (_, total_row) in enumerate(total_docentes.iterrows(), start=1):
        with multi_tabs[tab_idx]:
            rut = total_row['RUT']
            nombre = total_row['NOMBRE']
            brp_total_doc = total_row['BRP_TOTAL']

            st.markdown(f"**{nombre}** ({rut})")
            st.metric("BRP Total", fmt_clp(brp_total_doc))

            detalle = df_multi[(df_multi['RUT'] == rut) & (df_multi['TIPO_FILA'] == 'DETALLE')]

            cols_show = ['RBD']
            if 'ESCUELA' in detalle.columns:
                cols_show.append('ESCUELA')
            cols_show += ['HORAS_CONTRATO', 'RECONOCIMIENTO_MINEDUC', 'TRAMO_MINEDUC',
                          'PRIORITARIOS_MINEDUC', 'BRP_SEP', 'BRP_PIE', 'BRP_NORMAL', 'BRP_TOTAL']
            cols_show = [c for c in cols_show if c in detalle.columns]
            st.dataframe(detalle[cols_show], width='stretch', hide_index=True)

            if len(detalle) >= 2:
                chart_cols = st.columns(len(detalle))
                for i, (_, row) in enumerate(detalle.iterrows()):
                    rbd_label = row.get('ESCUELA', str(row['RBD'])) if 'ESCUELA' in detalle.columns else str(row['RBD'])
                    if not rbd_label:
                        rbd_label = str(row['RBD'])
                    with chart_cols[i]:
                        st.metric(f"RBD {row['RBD']}", fmt_clp(row['BRP_TOTAL']))
                        st.caption(rbd_label[:25])
                        st.caption(f"Hrs: {row['HORAS_CONTRATO']}")


def _find_col(df, *targets):
    """Busca una columna en el DataFrame por nombre parcial (case-insensitive)."""
    for t in targets:
        for col in df.columns:
            if t.lower() in col.lower():
                return col
    return None


@st.fragment
def show_data_explorer(df, key_prefix="expl"):
    """Explorador interactivo unificado con filtros, DAEM/CPEIP y drill-down."""
    st.markdown("---")
    st.markdown("##### 🔍 Explorador de Resultados")

    df_ex = df.copy()

    # Detectar columnas clave
    col_rbd = _find_col(df_ex, 'rbd')
    col_rut = _find_col(df_ex, 'rut')
    col_nombre = _find_col(df_ex, 'nombre_completo', 'nombre')

    # Forzar tipos string en columnas de texto mixto para evitar errores Arrow
    if col_rut:
        df_ex[col_rut] = df_ex[col_rut].astype(str)
    if col_rbd:
        df_ex[col_rbd] = df_ex[col_rbd].astype(str).str.replace(r'\.0$', '', regex=True)
    col_tipo_pago = _find_col(df_ex, 'tipo de pago', 'tipo_pago')
    has_mes = 'MES' in df_ex.columns

    if col_rbd:
        df_ex = add_school_names(df_ex, rbd_col=col_rbd)
    if col_rut:
        df_ex['RUT_FMT'] = df_ex[col_rut].apply(format_rut)

    brp_cols = [c for c in ['BRP_SEP', 'BRP_PIE', 'BRP_NORMAL', 'BRP_TOTAL'] if c in df_ex.columns]
    daem_cols = [c for c in ['TOTAL_DAEM_SEP', 'TOTAL_DAEM_PIE', 'TOTAL_DAEM_NORMAL'] if c in df_ex.columns]
    cpeip_cols = [c for c in ['TOTAL_CPEIP_SEP', 'TOTAL_CPEIP_PIE', 'TOTAL_CPEIP_NORMAL'] if c in df_ex.columns]

    # --- Filtros ---
    st.markdown("**Filtros**")
    filter_parts = []
    if has_mes:
        filter_parts.append('mes')
    if col_tipo_pago:
        filter_parts.append('tipo')
    if col_rbd:
        filter_parts.append('rbd')
    filter_parts.append('buscar')

    filter_cols = st.columns(len(filter_parts))
    idx = 0

    meses_sel = None
    if has_mes:
        with filter_cols[idx]:
            if 'MES_NUM' in df_ex.columns:
                mes_order = df_ex[['MES', 'MES_NUM']].drop_duplicates().sort_values('MES_NUM')
                meses_disponibles = mes_order['MES'].tolist()
            else:
                meses_disponibles = sorted(df_ex['MES'].unique().tolist())
            meses_sel = st.multiselect(
                "📅 Mes", meses_disponibles, default=meses_disponibles,
                key=f"{key_prefix}_mes"
            )
        idx += 1

    tipos_sel = None
    if col_tipo_pago:
        with filter_cols[idx]:
            tipos_pago = sorted(df_ex[col_tipo_pago].dropna().unique().tolist(), key=str)
            tipos_sel = st.multiselect(
                "💼 Tipo de Pago", tipos_pago, default=tipos_pago,
                key=f"{key_prefix}_tipo"
            )
        idx += 1

    rbd_sel = 'Todas'
    if col_rbd:
        with filter_cols[idx]:
            rbd_list = sorted(df_ex[col_rbd].dropna().unique().tolist(), key=str)
            escuela_map = {}
            if 'ESCUELA' in df_ex.columns:
                escuela_map = df_ex.groupby(col_rbd)['ESCUELA'].first().to_dict()

            def fmt_opt(r):
                if r == 'Todas':
                    return 'Todas'
                name = escuela_map.get(r, '')
                return f"{r} — {name}" if name else str(r)

            rbd_sel = st.selectbox(
                "🏫 Escuela", ['Todas'] + rbd_list,
                format_func=fmt_opt, key=f"{key_prefix}_rbd"
            )
        idx += 1

    with filter_cols[idx]:
        buscar = st.text_input("🔎 Buscar", key=f"{key_prefix}_buscar")

    # --- Aplicar filtros ---
    mask = pd.Series([True] * len(df_ex), index=df_ex.index)
    if meses_sel is not None and has_mes:
        mask &= df_ex['MES'].isin(meses_sel)
    if tipos_sel is not None and col_tipo_pago:
        mask &= df_ex[col_tipo_pago].isin(tipos_sel)
    if col_rbd and rbd_sel != 'Todas':
        mask &= df_ex[col_rbd] == rbd_sel
    if buscar.strip():
        term = buscar.strip().lower()
        text_mask = pd.Series([False] * len(df_ex), index=df_ex.index)
        if col_rut:
            text_mask |= df_ex[col_rut].astype(str).str.lower().str.contains(term, na=False)
        if col_nombre:
            text_mask |= df_ex[col_nombre].astype(str).str.lower().str.contains(term, na=False)
        mask &= text_mask

    df_f = df_ex[mask]

    if df_f.empty:
        st.warning("No hay datos con los filtros seleccionados.")
        return

    # --- Métricas: BRP + DAEM/CPEIP ---
    total_daem = sum(df_f[c].sum() for c in daem_cols) if daem_cols else 0
    total_cpeip = sum(df_f[c].sum() for c in cpeip_cols) if cpeip_cols else 0

    row1 = st.columns(4)
    row1[0].metric("Registros", f"{len(df_f):,}".replace(',', '.'))
    if col_rut:
        row1[1].metric("Docentes", f"{df_f[col_rut].nunique():,}".replace(',', '.'))
    if 'BRP_TOTAL' in df_f.columns:
        row1[2].metric("BRP Total", fmt_clp(df_f['BRP_TOTAL'].sum()))

    row2 = st.columns(5)
    if 'BRP_SEP' in df_f.columns:
        row2[0].metric("BRP SEP", fmt_clp(df_f['BRP_SEP'].sum()))
    if 'BRP_PIE' in df_f.columns:
        row2[1].metric("BRP PIE", fmt_clp(df_f['BRP_PIE'].sum()))
    if 'BRP_NORMAL' in df_f.columns:
        row2[2].metric("BRP Normal", fmt_clp(df_f['BRP_NORMAL'].sum()))
    if daem_cols:
        row2[3].metric("Paga DAEM", fmt_clp(total_daem))
    if cpeip_cols:
        row2[4].metric("Paga CPEIP", fmt_clp(total_cpeip))

    # --- Desglose por tipo de pago con DAEM/CPEIP ---
    if col_tipo_pago and col_tipo_pago in df_f.columns:
        st.markdown("**💼 Desglose por Tipo de Pago**")
        agg_tipo = {}
        if col_rut:
            agg_tipo['REGISTROS'] = (col_rut, 'count')
            agg_tipo['DOCENTES'] = (col_rut, 'nunique')
        for c in brp_cols:
            agg_tipo[c] = (c, 'sum')
        for c in daem_cols + cpeip_cols:
            agg_tipo[c] = (c, 'sum')
        df_tipo = df_f.groupby(col_tipo_pago).agg(**agg_tipo).reset_index()
        if daem_cols:
            df_tipo['TOTAL_DAEM'] = df_tipo[daem_cols].sum(axis=1)
        if cpeip_cols:
            df_tipo['TOTAL_CPEIP'] = df_tipo[cpeip_cols].sum(axis=1)

        show_tipo = [col_tipo_pago]
        if 'REGISTROS' in df_tipo.columns:
            show_tipo.append('REGISTROS')
        if 'DOCENTES' in df_tipo.columns:
            show_tipo.append('DOCENTES')
        show_tipo += brp_cols
        if 'TOTAL_DAEM' in df_tipo.columns:
            show_tipo.append('TOTAL_DAEM')
        if 'TOTAL_CPEIP' in df_tipo.columns:
            show_tipo.append('TOTAL_CPEIP')
        df_tipo_display = df_tipo[show_tipo].copy()
        df_tipo_display = append_totals_row(df_tipo_display, label_col=col_tipo_pago, label='TOTAL',
                                            skip_cols={col_tipo_pago})
        st.dataframe(
            df_tipo_display.style.format(format_money_cols(df_tipo_display, exclude_cols={col_tipo_pago})),
            width='stretch', hide_index=True,
        )
        add_table_downloads(df_tipo_display, 'desglose_tipo_pago', f'{key_prefix}_tipo')

    # --- Resumen por escuela con DAEM/CPEIP ---
    if col_rbd and rbd_sel == 'Todas':
        st.markdown("**🏫 Resumen por Escuela**")
        agg_esc = {}
        if col_rut:
            agg_esc['DOCENTES'] = (col_rut, 'nunique')
        for c in brp_cols:
            agg_esc[c] = (c, 'sum')
        for c in daem_cols + cpeip_cols:
            agg_esc[c] = (c, 'sum')
        df_esc = df_f.groupby(col_rbd).agg(**agg_esc).reset_index()
        df_esc = add_school_names(df_esc, rbd_col=col_rbd)
        if daem_cols:
            df_esc['TOTAL_DAEM'] = df_esc[daem_cols].sum(axis=1)
        if cpeip_cols:
            df_esc['TOTAL_CPEIP'] = df_esc[cpeip_cols].sum(axis=1)

        ordered = [col_rbd]
        if 'ESCUELA' in df_esc.columns:
            ordered.append('ESCUELA')
        if 'DOCENTES' in df_esc.columns:
            ordered.append('DOCENTES')
        ordered += brp_cols
        if 'TOTAL_DAEM' in df_esc.columns:
            ordered.append('TOTAL_DAEM')
        if 'TOTAL_CPEIP' in df_esc.columns:
            ordered.append('TOTAL_CPEIP')
        df_esc = df_esc[[c for c in ordered if c in df_esc.columns]]
        # Contar docentes únicos para la fila total
        docentes_unique = df_f[col_rut].nunique() if col_rut else None
        df_esc = append_totals_row(df_esc, label_col=col_rbd, label='TOTAL',
                                   skip_cols={'ESCUELA'},
                                   docentes_col='DOCENTES' if 'DOCENTES' in df_esc.columns else None,
                                   docentes_value=docentes_unique)
        st.dataframe(
            df_esc.style.format(format_money_cols(df_esc, exclude_cols={col_rbd, 'ESCUELA'})),
            width='stretch', hide_index=True,
        )
        add_table_downloads(df_esc, 'resumen_por_escuela', f'{key_prefix}_esc')

    # --- Tabla de docentes ---
    st.markdown("**👥 Docentes**")
    show_cols = []
    if has_mes:
        show_cols.append('MES')
    if 'RUT_FMT' in df_f.columns:
        show_cols.append('RUT_FMT')
    if col_nombre:
        show_cols.append(col_nombre)
    if col_tipo_pago:
        show_cols.append(col_tipo_pago)
    show_cols += [c for c in ['HORAS_SEP', 'HORAS_PIE', 'HORAS_SN'] if c in df_f.columns]
    show_cols += brp_cols
    if daem_cols:
        show_cols += daem_cols
    if cpeip_cols:
        show_cols += cpeip_cols
    show_cols = [c for c in show_cols if c in df_f.columns]

    df_docentes_display = df_f[show_cols].rename(columns={'RUT_FMT': 'RUT'}).copy()
    df_docentes_display = append_totals_row(df_docentes_display, label_col='RUT' if 'RUT' in df_docentes_display.columns else show_cols[0], label='TOTAL')
    st.dataframe(
        df_docentes_display.style.format(format_money_cols(df_docentes_display)),
        width='stretch', hide_index=True,
    )
    add_table_downloads(df_docentes_display, 'detalle_docentes', f'{key_prefix}_docs')

    # --- Detalle docente ---
    df_f = df_f.reset_index(drop=True)
    _col_nombre = col_nombre  # capture for lambda
    _has_mes = has_mes
    labels = df_f.apply(
        lambda r: (
            f"{r.get('RUT_FMT', '')} — "
            f"{r.get(_col_nombre, '') if _col_nombre else ''}"
            + (f" ({r.get('MES', '')})" if _has_mes else "")
        ),
        axis=1
    ).tolist()

    if labels:
        sel_idx = st.selectbox(
            "👤 Detalle Docente", range(len(labels)),
            format_func=lambda i: labels[i], key=f"{key_prefix}_doc"
        )
        teacher = df_f.iloc[sel_idx]

        mc = st.columns(min(len(brp_cols), 4))
        for i, col in enumerate(brp_cols):
            with mc[i]:
                st.metric(col.replace('BRP_', ''), fmt_clp(teacher.get(col, 0)))

        if daem_cols or cpeip_cols:
            dc = st.columns(2)
            if daem_cols:
                dc[0].metric("Paga DAEM", fmt_clp(sum(teacher.get(c, 0) for c in daem_cols)))
            if cpeip_cols:
                dc[1].metric("Paga CPEIP", fmt_clp(sum(teacher.get(c, 0) for c in cpeip_cols)))

        hr_cols = [c for c in ['HORAS_SEP', 'HORAS_PIE', 'HORAS_SN'] if c in df_f.columns]
        if hr_cols:
            hc = st.columns(len(hr_cols))
            for i, col in enumerate(hr_cols):
                with hc[i]:
                    st.metric(col.replace('HORAS_', 'Hrs '), teacher.get(col, 0))

        with st.expander("Ver todos los campos"):
            detail = {k: v for k, v in teacher.items() if k != 'RUT_FMT'}
            df_det = pd.DataFrame([detail])
            if col_rut and col_rut in df_det.columns:
                df_det[col_rut] = df_det[col_rut].apply(format_rut)
            st.dataframe(
                df_det.style.format(format_money_cols(df_det)),
                width='stretch', hide_index=True,
            )


def show_sidebar_charts():
    """Muestra charts dinámicos en el sidebar basados en resultados del procesamiento."""
    if 'last_brp_result' not in st.session_state:
        return

    data = st.session_state['last_brp_result']
    mes = data.get('mes', '')
    brp_sep = data.get('brp_sep', 0)
    brp_pie = data.get('brp_pie', 0)
    brp_normal = data.get('brp_normal', 0)
    brp_total = data.get('brp_total', 0)
    daem_total = data.get('daem_total', 0)
    cpeip_total = data.get('cpeip_total', 0)

    if brp_total <= 0:
        return

    import plotly.graph_objects as go

    st.markdown(f"**📅 Mes: {mes}**")
    st.metric("BRP Total", fmt_clp(brp_total))

    # Mini gráfico de torta SEP/PIE/Normal
    fig = go.Figure(data=[go.Pie(
        labels=['SEP', 'PIE', 'Normal'],
        values=[brp_sep, brp_pie, brp_normal],
        marker_colors=['#3b82f6', '#10b981', '#f59e0b'],
        textinfo='percent',
        hole=0.4,
    )])
    fig.update_layout(
        height=200, margin=dict(l=10, r=10, t=25, b=10),
        title_text='SEP / PIE / Normal', title_font_size=12,
        showlegend=True,
        legend=dict(font=dict(size=10), orientation='h', y=-0.1),
        separators=',.'
    )
    st.plotly_chart(fig, width='stretch')

    # DAEM vs CPEIP
    if daem_total > 0 or cpeip_total > 0:
        fig2 = go.Figure(data=[go.Bar(
            x=['DAEM', 'CPEIP'],
            y=[daem_total, cpeip_total],
            marker_color=['#6366f1', '#ec4899'],
            text=[fmt_clp(daem_total), fmt_clp(cpeip_total)],
            textposition='inside',
        )])
        fig2.update_layout(
            height=200, margin=dict(l=10, r=10, t=25, b=10),
            title_text='DAEM vs CPEIP', title_font_size=12,
            showlegend=False, yaxis_visible=False,
            separators=',.'
        )
        st.plotly_chart(fig2, width='stretch')

    # Detalle por tipo
    st.caption(f"SEP: {fmt_clp(brp_sep)}")
    st.caption(f"PIE: {fmt_clp(brp_pie)}")
    st.caption(f"Normal: {fmt_clp(brp_normal)}")

    # Info REM si existe
    if 'last_rem_result' in st.session_state:
        rem = st.session_state['last_rem_result']
        st.markdown("---")
        st.markdown("**📋 REM**")
        st.caption(f"Personas: {rem.get('personas', 0)}")
        st.caption(f"Horas totales: {rem.get('total_horas', 0)}")
        exceden = rem.get('exceden', 0)
        if exceden > 0:
            st.badge(f"{exceden} exceden 44hrs", color="red")
        else:
            st.badge("Todos dentro del límite", color="green")


def card_start(title: str, icon: str = ""):
    """Inicia una tarjeta (HTML-sanitized)."""
    safe_title = _sanitize_html(title)
    safe_icon = _sanitize_html(icon)
    st.markdown(f"""
    <div class="custom-card">
        <div class="card-header">{safe_icon} {safe_title}</div>
        <div class="card-content">
    """, unsafe_allow_html=True)


def card_end():
    """Cierra una tarjeta."""
    st.markdown('</div></div>', unsafe_allow_html=True)


def show_tutorial(steps: list):
    """Muestra pasos del tutorial con componentes nativos Streamlit."""
    cols = st.columns(min(len(steps), 3))
    for i, (title, desc) in enumerate(steps):
        with cols[i % min(len(steps), 3)]:
            st.info(f"**{i+1}. {title}**\n\n{desc}")


def check_sheets(file, required: list) -> tuple:
    """Valida hojas del Excel."""
    try:
        xlsx = pd.ExcelFile(file)
        missing = [s for s in required if s not in xlsx.sheet_names]
        return len(missing) == 0, missing
    except Exception as e:
        return False, [str(e)]


def to_excel_buffer(df: pd.DataFrame) -> BytesIO:
    """Convierte DataFrame a buffer para descarga con formato profesional."""
    return to_styled_excel(df)


def to_styled_excel(df: pd.DataFrame, sheet_name='Datos', title=None) -> BytesIO:
    """Genera Excel con formato profesional: CLP, headers con estilo, totales bold.

    Args:
        df: DataFrame a exportar.
        sheet_name: nombre de la hoja.
        title: título opcional en la primera fila.
    Returns:
        BytesIO con el Excel generado.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    start_row = 1
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        if title:
            start_row = 3
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row - 1)
        ws = writer.sheets[sheet_name]

        # Título
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            cell_title = ws.cell(row=1, column=1, value=title)
            cell_title.font = Font(bold=True, size=13, color='1E293B')
            cell_title.alignment = Alignment(horizontal='center')

        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        total_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )
        clp_format = '#,##0'

        # Detectar columnas monetarias
        money_col_indices = set()
        for i, col in enumerate(df.columns):
            if col not in NON_MONEY_COLS and df[col].dtype in ('float64', 'int64', 'float32', 'int32'):
                money_col_indices.add(i + 1)  # openpyxl es 1-indexed

        # Header row
        header_row = start_row
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for row_idx in range(start_row + 1, start_row + 1 + len(df)):
            is_total = False
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if str(cell.value).upper().startswith('TOTAL'):
                    is_total = True
                    break

            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if is_total:
                    cell.font = total_font
                    cell.fill = total_fill
                if col_idx in money_col_indices:
                    cell.number_format = clp_format
                    cell.alignment = Alignment(horizontal='right')

        # Auto-ancho columnas
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(df.columns[col_idx - 1]))
            for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(df),
                                     min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        val_len = len(str(cell.value))
                        if col_idx in money_col_indices and isinstance(cell.value, (int, float)):
                            val_len = len(f'{int(cell.value):,}') + 1
                        max_len = max(max_len, val_len)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

        # Freeze panes
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    buf.seek(0)
    return buf


def add_table_downloads(df, label, key, include_pdf=True, excel_title='auto'):
    """Agrega botones de descarga Excel + PDF debajo de una tabla.

    Args:
        df: DataFrame a descargar.
        label: nombre base del archivo (ej: 'resumen_por_escuela').
        key: clave única para los botones streamlit.
        include_pdf: si incluir botón PDF.
        excel_title: título del Excel. 'auto' genera desde label; None omite título.
    """
    if excel_title == 'auto':
        excel_title = label.replace('_', ' ').title()
    cols = st.columns(3 if include_pdf else 2)
    with cols[0]:
        excel_bytes = to_styled_excel(df, title=excel_title)
        st.download_button(
            "📥 Excel", data=excel_bytes,
            file_name=f"{label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_{key}_xlsx",
        )
    with cols[1]:
        csv_bytes = df.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            "📥 CSV", data=csv_bytes,
            file_name=f"{label}.csv",
            mime="text/csv",
            key=f"dl_{key}_csv",
        )
    if include_pdf:
        with cols[2]:
            try:
                pdf_bytes = generate_pdf_from_df(df, title=label.replace('_', ' ').title())
                st.download_button(
                    "📥 PDF", data=pdf_bytes,
                    file_name=f"{label}.pdf",
                    mime="application/pdf",
                    key=f"dl_{key}_pdf",
                )
            except Exception:
                st.caption("PDF no disponible")


def _cleanup_temp_files(*paths):
    """Securely remove temporary files containing sensitive salary data."""
    import os
    for p in paths:
        try:
            if p and Path(p).exists():
                os.unlink(p)
        except OSError:
            pass


def process_files(processor, inputs: list):
    """Procesa archivos con barra de progreso. Cleans up temp files after use."""
    progress = st.progress(0)
    status = st.empty()

    def callback(val, msg):
        progress.progress(val / 100)
        status.markdown(f"**{msg}**")

    paths = []
    out_path = None
    try:
        # Crear archivos temporales (preserva extensión original para detección CSV/Excel)
        for f in inputs:
            original_ext = Path(f.name).suffix or '.xlsx'
            tmp = tempfile.NamedTemporaryFile(suffix=original_ext, delete=False)
            tmp.write(f.getvalue())
            paths.append(Path(tmp.name))
            tmp.close()

        out_tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        out_path = Path(out_tmp.name)
        out_tmp.close()

        # Ejecutar procesador segun cantidad de archivos
        if len(paths) == 1:
            processor.process_file(paths[0], out_path, callback)
        elif len(paths) == 2:
            processor.process_file(paths[0], paths[1], out_path, callback)
        elif len(paths) == 3:
            processor.process_file(
                web_sostenedor_path=paths[0],
                sep_procesado_path=paths[1],
                pie_procesado_path=paths[2],
                output_path=out_path,
                progress_callback=callback
            )

        df = pd.read_excel(out_path, engine='openpyxl')
        progress.progress(100)
        status.markdown("**Completado**")
        return df, None

    except Exception as e:
        return None, format_user_error(e)
    finally:
        # Always clean up temp files to prevent sensitive data leaks
        _cleanup_temp_files(*paths, out_path)


# ============================================================================
# PESTAÑAS
# ============================================================================

def tab_sep_pie():
    """Pestaña de procesamiento SEP/PIE/EIB."""

    with st.expander("📖 ¿Cómo usar esta herramienta?", expanded=False):
        show_tutorial([
            ("Selecciona el tipo", "Elige SEP, PIE-NORMAL o EIB según el archivo a procesar."),
            ("Sube el archivo", "SEP/PIE: Excel con hojas HORAS y TOTAL. EIB: Excel con Hoja1 y columna Jornada."),
            ("Procesa", "Haz clic en el botón verde y espera."),
            ("Descarga", "Guarda el archivo procesado en tu computador.")
        ])

    st.markdown("---")

    # Configuración
    col1, col2 = st.columns([1, 2])
    with col1:
        modo = st.selectbox(
            "📋 Tipo de procesamiento",
            ["SEP", "PIE-NORMAL", "EIB"],
            help="SEP: Subvención Escolar Preferencial\nPIE-NORMAL: Programa de Integración + Subvención Normal\nEIB: Educación Intercultural Bilingüe"
        )

    st.markdown("")

    # Archivo
    st.markdown("##### 📁 Archivo de Entrada")
    if modo == "EIB":
        st.caption("El archivo debe contener una hoja con datos y columna **Jornada**")
    else:
        st.caption("El archivo debe contener las hojas **HORAS** y **TOTAL**")
    archivo = st.file_uploader(
        "Arrastra o selecciona un archivo (Excel o CSV)",
        type=['xlsx', 'xls', 'csv'],
        key="sep_file"
    )

    if archivo:
        is_csv = archivo.name.lower().endswith('.csv')

        if modo == "EIB":
            # EIB: CSV o Excel con al menos 1 hoja
            if is_csv:
                st.success(f"✅ **{archivo.name}** - Archivo CSV válido")
            else:
                try:
                    xlsx = pd.ExcelFile(archivo)
                    if not xlsx.sheet_names:
                        st.error("❌ **Archivo incorrecto** - No contiene hojas")
                        return
                    st.success(f"✅ **{archivo.name}** - Archivo válido (hoja: {xlsx.sheet_names[0]})")
                except Exception as e:
                    st.error(f"❌ **Error al leer archivo:** {e}")
                    return
        else:
            # SEP/PIE requieren hojas HORAS+TOTAL (no disponible en CSV)
            if is_csv:
                st.error("❌ **SEP/PIE requieren un archivo Excel** con hojas HORAS y TOTAL. Los archivos CSV no tienen hojas.")
                return
            ok, missing = check_sheets(archivo, ['HORAS', 'TOTAL'])
            if not ok:
                st.error(f"❌ **Archivo incorrecto** - Faltan hojas: {', '.join(missing)}")
                return
            st.success(f"✅ **{archivo.name}** - Archivo válido")

        st.markdown("")

        if st.button("▶️  PROCESAR ARCHIVO", key="btn_sep", width='stretch'):
            if modo == "EIB":
                processor = EIBProcessor()
            elif modo == "SEP":
                processor = SEPProcessor()
            else:
                processor = PIEProcessor()
            df, error = process_files(processor, [archivo])

            if error:
                st.error(f"❌ **Error:** {error}")
            else:
                success_box(f"Se procesaron **{len(df)}** registros correctamente")
                st.toast(f"{modo} procesado: {len(df)} registros", icon="✅")

                c1, c2 = st.columns(2)
                with c1:
                    st.metric("Registros", f"{len(df)}")
                with c2:
                    st.metric("Columnas", f"{len(df.columns)}")

                st.dataframe(df.head(10), width='stretch', hide_index=True)

                nombre = f"{Path(archivo.name).stem}_procesado.xlsx"

                _, col_dl, _ = st.columns([1, 2, 1])
                with col_dl:
                    st.download_button(
                        "📥  DESCARGAR RESULTADO",
                        data=to_excel_buffer(df),
                        file_name=nombre,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )


def tab_brp():
    """Pestaña de distribución BRP."""
    
    with st.expander("📖 ¿Cómo usar esta herramienta?", expanded=False):
        show_tutorial([
            ("Procesa primero", "Ve a la pestaña SEP/PIE y procesa ambos archivos por separado."),
            ("Carga los archivos", "Arrastra los 3 archivos - se detectan automáticamente por nombre."),
            ("Valida (opcional)", "Revisa que los archivos sean del mismo mes."),
            ("Distribuye", "El sistema calcula BRP para SEP, PIE y NORMAL.")
        ])
    
    info_box("Los archivos se detectan por nombre: <b>web*</b> → MINEDUC, <b>*sep*</b> → SEP, <b>*pie*/*sn*/*normal*</b> → PIE/Normal")
    
    st.markdown("---")
    
    # Estado para los archivos
    if 'brp_files' not in st.session_state:
        st.session_state.brp_files = {'web': None, 'sep': None, 'pie': None}
    
    # Uploader múltiple
    st.markdown("##### 📥 Cargar Archivos (arrástralos todos juntos)")
    uploaded_files = st.file_uploader(
        "Arrastra los 3 archivos (Excel o CSV)",
        type=['xlsx', 'xls', 'csv'],
        accept_multiple_files=True,
        key="brp_multi_upload"
    )
    
    # Auto-detectar y asignar archivos
    for f in uploaded_files:
        name_lower = f.name.lower()
        if name_lower.startswith('web'):
            st.session_state.brp_files['web'] = f
        elif name_lower.startswith('sep') or 'sep' in name_lower:
            st.session_state.brp_files['sep'] = f
        elif name_lower.startswith('sn') or 'pie' in name_lower or 'normal' in name_lower:
            st.session_state.brp_files['pie'] = f
    
    # Mostrar estado de archivos
    st.markdown("##### 📋 Archivos detectados")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**📋 MINEDUC (web_sostenedor)**")
        f_web = st.session_state.brp_files['web']
        if f_web:
            st.success(f"✓ {f_web.name}")
        else:
            st.warning("⬜ No detectado (nombre debe empezar con 'web')")
    
    with col2:
        st.markdown("**📊 SEP procesado**")
        f_sep = st.session_state.brp_files['sep']
        if f_sep:
            st.success(f"✓ {f_sep.name}")
        else:
            st.warning("⬜ No detectado (nombre debe contener 'sep')")
    
    with col3:
        st.markdown("**📊 PIE/Normal procesado**")
        f_pie = st.session_state.brp_files['pie']
        if f_pie:
            st.success(f"✓ {f_pie.name}")
        else:
            st.warning("⬜ No detectado (nombre debe contener 'sn', 'pie' o 'normal')")
    
    # Botón para limpiar
    if st.button("🔄 Limpiar archivos", key="btn_clear_brp"):
        st.session_state.brp_files = {'web': None, 'sep': None, 'pie': None}
        st.rerun()
    
    # Verificar que están todos
    if not all([f_web, f_sep, f_pie]):
        warning_box("Carga los **3 archivos** para continuar")
        return

    # Detectar meses del web sostenedor
    brp_month_filter = None
    try:
        tmp_web_detect = tempfile.NamedTemporaryFile(
            suffix=Path(f_web.name).suffix or '.xlsx', delete=False
        )
        tmp_web_detect.write(f_web.getvalue())
        tmp_web_detect.close()
        web_months = BRPProcessor.detect_web_months(Path(tmp_web_detect.name))
        import os
        os.unlink(tmp_web_detect.name)
    except Exception:
        web_months = None

    if web_months and len(web_months) > 1:
        st.markdown("---")
        st.markdown("##### 📅 Seleccionar Mes del Web Sostenedor")
        st.info(f"El archivo web contiene **{len(web_months)} meses**. Selecciona cuál procesar.")
        # web_months es ['01', '02', ...] - mostrar nombres bonitos
        month_options = {MESES_NUM_TO_NAME.get(m, m): m for m in web_months}
        mes_display = st.selectbox(
            "Mes a procesar",
            options=list(month_options.keys()),
            key="brp_web_month_select"
        )
        if mes_display:
            brp_month_filter = month_options[mes_display]

    # Validación de archivos del mismo mes
    st.markdown("---")
    st.markdown("##### ⚙️ Opciones")

    col_opt1, col_opt2 = st.columns(2)
    with col_opt1:
        solo_validar = st.checkbox("🔍 Solo validar (sin procesar)", value=False,
                                   help="Revisa los archivos sin generar el resultado final")
    
    # Detectar mes de los archivos por nombre
    meses_detectados = []
    for f, tipo in [(f_sep, 'SEP'), (f_pie, 'PIE')]:
        name = f.name.lower()
        for mes in ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 
                    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']:
            if mes in name:
                meses_detectados.append((tipo, mes.capitalize()))
                break
    
    if len(meses_detectados) >= 2:
        mes_sep = next((m for t, m in meses_detectados if t == 'SEP'), None)
        mes_pie = next((m for t, m in meses_detectados if t == 'PIE'), None)
        if mes_sep and mes_pie:
            if mes_sep == mes_pie:
                st.success(f"✅ Ambos archivos son de **{mes_sep}**")
            else:
                st.error(f"⚠️ **Meses diferentes:** SEP={mes_sep}, PIE={mes_pie}")
    
    st.markdown("---")
    
    btn_text = "🔍 VALIDAR ARCHIVOS" if solo_validar else "📊 DISTRIBUIR BRP"
    
    if st.button(btn_text, key="btn_brp", width='stretch'):
        processor = BRPProcessor()

        progress = st.progress(0)
        status = st.empty()

        def callback(val, msg):
            progress.progress(val / 100)
            status.markdown(f"**⏳ {msg}**")

        try:
            # Crear archivos temporales
            paths = []
            for f in [f_web, f_sep, f_pie]:
                tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                tmp.write(f.getvalue())
                paths.append(Path(tmp.name))
                tmp.close()

            out_tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            out_path = Path(out_tmp.name)
            out_tmp.close()

            # Procesar
            processor.process_file(
                web_sostenedor_path=paths[0],
                sep_procesado_path=paths[1],
                pie_procesado_path=paths[2],
                output_path=out_path,
                progress_callback=callback,
                month_filter=brp_month_filter,
            )

            # Mostrar alertas de columnas
            show_column_alerts(processor.get_column_alerts())

            # Leer resultado
            df = pd.read_excel(out_path, sheet_name='BRP_DISTRIBUIDO', engine='openpyxl')
            progress.progress(100)

            # Métricas principales
            brp_sep = df['BRP_SEP'].sum() if 'BRP_SEP' in df.columns else 0
            brp_pie = df['BRP_PIE'].sum() if 'BRP_PIE' in df.columns else 0
            brp_normal = df['BRP_NORMAL'].sum() if 'BRP_NORMAL' in df.columns else 0
            total = brp_sep + brp_pie + brp_normal

            if solo_validar:
                status.markdown("**✅ ¡Validación completada!**")
                st.info("📋 Modo validación: revisa los datos antes de procesar definitivamente")
            else:
                status.markdown("**✅ ¡Completado!**")
                success_box("¡Distribución de BRP completada!")
                st.toast(f"BRP distribuido: {fmt_clp(total)}", icon="💰")

            # Leer bytes del Excel y hojas adicionales
            with open(out_path, 'rb') as f:
                excel_bytes = f.read()

            df_rbd = None
            try:
                df_rbd = pd.read_excel(out_path, sheet_name='RESUMEN_POR_RBD', engine='openpyxl')
            except (ValueError, KeyError):
                pass

            df_revision = None
            try:
                df_revision = pd.read_excel(out_path, sheet_name='REVISAR', engine='openpyxl')
            except (ValueError, KeyError):
                pass

            # Guardar en session_state para sidebar charts
            brp_cols_list = df.columns
            daem_t = sum(df[f'TOTAL_DAEM_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL'] if f'TOTAL_DAEM_{s}' in brp_cols_list)
            cpeip_t = sum(df[f'TOTAL_CPEIP_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL'] if f'TOTAL_CPEIP_{s}' in brp_cols_list)
            st.session_state['last_brp_result'] = {
                'mes': 'BRP',
                'brp_sep': brp_sep, 'brp_pie': brp_pie,
                'brp_normal': brp_normal, 'brp_total': total,
                'daem_total': daem_t, 'cpeip_total': cpeip_t,
            }

            # Cachear resultados
            st.session_state['brp_tab_result'] = {
                'df': df,
                'excel_bytes': excel_bytes,
                'df_rbd': df_rbd,
                'df_revision': df_revision,
                'brp_sep': brp_sep, 'brp_pie': brp_pie,
                'brp_normal': brp_normal, 'total': total,
                'solo_validar': solo_validar,
            }

        except Exception as e:
            st.error(f"❌ **Error:** {format_user_error(e)}")
            with st.expander("Ver detalles técnicos"):
                st.code(str(e))

    # Mostrar resultados cacheados (persisten entre reruns)
    if 'brp_tab_result' in st.session_state:
        cached = st.session_state['brp_tab_result']
        df = cached['df']
        brp_sep = cached['brp_sep']
        brp_pie = cached['brp_pie']
        brp_normal = cached['brp_normal']
        total = cached['total']

        st.markdown("---")

        # Métricas principales siempre visibles
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("SEP", fmt_clp(brp_sep))
            if total > 0:
                st.badge(f"{100*brp_sep/total:.1f}%", color="blue")
        with c2:
            st.metric("PIE", fmt_clp(brp_pie))
            if total > 0:
                st.badge(f"{100*brp_pie/total:.1f}%", color="green")
        with c3:
            st.metric("NORMAL", fmt_clp(brp_normal))
            if total > 0:
                st.badge(f"{100*brp_normal/total:.1f}%", color="orange")
        with c4:
            st.metric("TOTAL", fmt_clp(total))
            st.badge("BRP Total", color="violet")

        res_tabs = st.tabs(["📈 Resumen", "📊 Gráficos", "🔍 Detalle", "📥 Descarga"])

        with res_tabs[0]:
            if total > 0:
                show_desglose_daem_cpeip(df)

        with res_tabs[1]:
            if total > 0:
                import plotly.express as px

                st.markdown("##### 📊 Distribución Visual")

                col_chart1, col_chart2 = st.columns(2)

                with col_chart1:
                    df_pie_chart = pd.DataFrame({
                        'Tipo': ['SEP', 'PIE', 'NORMAL'],
                        'Monto': [brp_sep, brp_pie, brp_normal]
                    })

                    fig = px.pie(df_pie_chart, values='Monto', names='Tipo',
                                 title='Distribución por Subvención',
                                 color_discrete_sequence=['#3b82f6', '#10b981', '#f59e0b'])
                    fig.update_traces(textposition='inside', textinfo='percent+label')
                    fig.update_layout(height=350, showlegend=True, separators=',.')
                    st.plotly_chart(fig, width='stretch')

                with col_chart2:
                    recon_total = (df['BRP_RECONOCIMIENTO_SEP'].sum() +
                                   df['BRP_RECONOCIMIENTO_PIE'].sum() +
                                   df['BRP_RECONOCIMIENTO_NORMAL'].sum())
                    tramo_total = (df['BRP_TRAMO_SEP'].sum() +
                                   df['BRP_TRAMO_PIE'].sum() +
                                   df['BRP_TRAMO_NORMAL'].sum())
                    prior_total = (df['CPEIP_PRIOR_SEP'].sum() +
                                   df['CPEIP_PRIOR_PIE'].sum() +
                                   df['CPEIP_PRIOR_NORMAL'].sum()) if 'CPEIP_PRIOR_SEP' in df.columns else 0

                    conceptos = ['Reconocimiento', 'Tramo']
                    montos = [recon_total, tramo_total]
                    colores = ['#8b5cf6', '#ec4899']
                    if prior_total > 0:
                        conceptos.append('Prioritarios')
                        montos.append(prior_total)
                        colores.append('#f97316')

                    df_concepto = pd.DataFrame({
                        'Concepto': conceptos,
                        'Monto': montos
                    })

                    fig2 = px.pie(df_concepto, values='Monto', names='Concepto',
                                  title='Distribución por Concepto',
                                  color_discrete_sequence=colores)
                    fig2.update_traces(textposition='inside', textinfo='percent+label')
                    fig2.update_layout(height=350, showlegend=True, separators=',.')
                    st.plotly_chart(fig2, width='stretch')

                # Gráficos por escuela si hay RBD
                df_rbd = cached.get('df_rbd')
                if df_rbd is not None and not df_rbd.empty:
                    st.markdown("##### 🏫 Distribución por Escuela")
                    show_charts_by_school(df_rbd)
            else:
                st.info("No hay datos para graficar.")

        with res_tabs[2]:
            # Resumen por RBD
            df_rbd = cached.get('df_rbd')
            if df_rbd is not None and not df_rbd.empty:
                st.markdown("##### 🏫 Resumen por Establecimiento")
                df_rbd_display = add_school_names(df_rbd).copy()
                # Agregar fila de totales
                money_cols_rbd = [c for c in df_rbd_display.columns if c not in ('RBD', 'ESCUELA') and df_rbd_display[c].dtype in ['float64', 'int64']]
                totals_rbd = {c: df_rbd_display[c].sum() for c in money_cols_rbd}
                totals_rbd['RBD'] = 'TOTAL'
                if 'ESCUELA' in df_rbd_display.columns:
                    totals_rbd['ESCUELA'] = ''
                df_rbd_display = pd.concat([df_rbd_display, pd.DataFrame([totals_rbd])], ignore_index=True)
                st.dataframe(
                    df_rbd_display.style.format({c: fmt_clp_style for c in money_cols_rbd}),
                    width='stretch', hide_index=True,
                )

            # Multi-Establecimiento
            show_multi_establishment(cached['excel_bytes'])

            # Detalle por concepto con tabla cruzada DAEM/CPEIP + totales
            st.markdown("---")
            st.markdown("##### 📊 Detalle por concepto (DAEM/CPEIP)")
            detalle_rows = []
            for subv in ['SEP', 'PIE', 'NORMAL']:
                daem_r = df[f'DAEM_RECON_{subv}'].sum() if f'DAEM_RECON_{subv}' in df.columns else 0
                cpeip_r = df[f'CPEIP_RECON_{subv}'].sum() if f'CPEIP_RECON_{subv}' in df.columns else 0
                detalle_rows.append({
                    'Concepto': 'Reconocimiento', 'Subvención': subv,
                    'DAEM ($)': int(daem_r), 'CPEIP ($)': int(cpeip_r),
                    'Total ($)': int(daem_r + cpeip_r)
                })
                daem_t_val = df[f'DAEM_TRAMO_{subv}'].sum() if f'DAEM_TRAMO_{subv}' in df.columns else 0
                cpeip_t_val = df[f'CPEIP_TRAMO_{subv}'].sum() if f'CPEIP_TRAMO_{subv}' in df.columns else 0
                detalle_rows.append({
                    'Concepto': 'Tramo', 'Subvención': subv,
                    'DAEM ($)': int(daem_t_val), 'CPEIP ($)': int(cpeip_t_val),
                    'Total ($)': int(daem_t_val + cpeip_t_val)
                })
                cpeip_p = df[f'CPEIP_PRIOR_{subv}'].sum() if f'CPEIP_PRIOR_{subv}' in df.columns else 0
                detalle_rows.append({
                    'Concepto': 'Prioritarios', 'Subvención': subv,
                    'DAEM ($)': 0, 'CPEIP ($)': int(cpeip_p),
                    'Total ($)': int(cpeip_p)
                })

            df_detalle = pd.DataFrame(detalle_rows)
            # Agregar fila de totales
            totals_row = {
                'Concepto': 'TOTAL', 'Subvención': '',
                'DAEM ($)': df_detalle['DAEM ($)'].sum(),
                'CPEIP ($)': df_detalle['CPEIP ($)'].sum(),
                'Total ($)': df_detalle['Total ($)'].sum(),
            }
            df_detalle = pd.concat([df_detalle, pd.DataFrame([totals_row])], ignore_index=True)
            st.dataframe(
                df_detalle.style.format({
                    'DAEM ($)': fmt_clp_style, 'CPEIP ($)': fmt_clp_style, 'Total ($)': fmt_clp_style
                }),
                width='stretch', hide_index=True,
            )

            # Casos para revisión
            df_revision = cached.get('df_revision')
            if df_revision is not None and not df_revision.empty:
                st.markdown("---")
                st.markdown("##### ⚠️ Casos para Revisión")

                exceden_44 = df_revision[df_revision['MOTIVO'] == 'EXCEDE 44 HORAS']
                sin_liquidacion = df_revision[df_revision['MOTIVO'] == 'SIN LIQUIDACIÓN']

                col1, col2 = st.columns(2)
                with col1:
                    if len(exceden_44) > 0:
                        st.badge(f"{len(exceden_44)} exceden 44hrs", icon="⚠️", color="orange")
                with col2:
                    if len(sin_liquidacion) > 0:
                        st.badge(f"{len(sin_liquidacion)} sin liquidación", icon="ℹ️", color="blue")

                with st.expander("👀 Ver casos a revisar"):
                    show_revision_table(df_revision)
            else:
                st.success("✅ No hay casos pendientes de revisión")

        with res_tabs[3]:
            if not cached.get('solo_validar'):
                st.markdown("##### 📁 Archivo generado")
                st.markdown("""
                El archivo contiene **5 hojas**:
                - `BRP_DISTRIBUIDO` → Datos completos con montos MINEDUC originales
                - `RESUMEN_POR_RBD` → Totales por establecimiento
                - `REVISAR` → Casos que requieren revisión
                - `RESUMEN_GENERAL` → Dashboard de resumen
                - `MULTI_ESTABLECIMIENTO` → Desglose de docentes en 2+ escuelas
                """)

                styled_brp_ind = _style_excel_workbook(cached['excel_bytes'])
                col_d1, col_d2 = st.columns(2)
                with col_d1:
                    st.download_button(
                        "📥 Excel Formateado",
                        data=styled_brp_ind,
                        file_name="brp_distribuido.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_brp_xlsx",
                    )
                with col_d2:
                    try:
                        df_brp_csv = pd.read_excel(BytesIO(cached['excel_bytes']), sheet_name='BRP_DISTRIBUIDO', engine='openpyxl')
                        csv_brp = df_brp_csv.to_csv(index=False).encode('utf-8-sig')
                        st.download_button(
                            "📥 CSV Detalle",
                            data=csv_brp,
                            file_name="brp_distribuido.csv",
                            mime="text/csv",
                            key="dl_brp_csv",
                        )
                    except Exception:
                        pass
            else:
                st.info("Modo validación: no se genera archivo para descargar.")


def tab_duplicados():
    """Pestaña de procesamiento de duplicados."""
    
    with st.expander("📖 ¿Cómo usar esta herramienta?", expanded=False):
        show_tutorial([
            ("Carga el archivo principal", "Debe tener una columna llamada DUPLICADOS."),
            ("Carga el complementario", "Archivo adicional para cruzar información."),
            ("Procesa", "El sistema consolida registros duplicados sumando valores.")
        ])
    
    st.markdown("""
    Esta herramienta consolida registros duplicados **sumando valores numéricos** 
    y manteniendo la **primera ocurrencia** de los datos textuales.
    """)
    
    st.markdown("---")
    
    # Archivos
    st.markdown("##### 📥 Archivos de Entrada")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.caption("📄 **Archivo Principal**")
        f1 = st.file_uploader("Principal", type=['xlsx', 'xls', 'csv'], key="dup_1", label_visibility="collapsed")
        if f1:
            st.success(f"✓ {f1.name}")
    
    with col2:
        st.caption("📄 **Archivo Complementario**")
        f2 = st.file_uploader("Complementario", type=['xlsx', 'xls', 'csv'], key="dup_2", label_visibility="collapsed")
        if f2:
            st.success(f"✓ {f2.name}")
    
    if not all([f1, f2]):
        warning_box("Carga **ambos archivos** para continuar")
        return
    
    st.markdown("")
    
    if st.button("🔄  PROCESAR DUPLICADOS", key="btn_dup", width='stretch'):
        processor = DuplicadosProcessor()
        df, error = process_files(processor, [f1, f2])
        
        if error:
            st.error(f"❌ **Error:** {error}")
        else:
            success_box(f"Se consolidaron **{len(df)}** registros")
            st.toast(f"Duplicados consolidados: {len(df)} registros", icon="🔄")

            nombre = f"{Path(f1.name).stem}_consolidado.xlsx"
            _, col_dl, _ = st.columns([1, 2, 1])
            with col_dl:
                st.download_button(
                    "📥  DESCARGAR RESULTADO",
                    data=to_excel_buffer(df),
                    file_name=nombre,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


def tab_todo_en_uno():
    """Pestaña de procesamiento integrado Todo en Uno."""

    with st.expander("📖 ¿Cómo usar esta herramienta?", expanded=False):
        show_tutorial([
            ("Carga los 3 archivos BRUTOS", "Arrastra SEP bruto, PIE bruto y web_sostenedor juntos."),
            ("Indica el mes", "Escribe el período en formato YYYY-MM (ej: 2024-01)."),
            ("Compara (opcional)", "Activa comparación si tienes meses anteriores guardados."),
            ("Procesa y descarga", "Obtén Excel procesado e Informe Word con auditoría.")
        ])

    info_box("Carga los <b>3 archivos BRUTOS</b> (SEP, PIE, web_sostenedor) y el sistema procesará todo automáticamente.")

    st.markdown("---")

    # Inicializar repositorio y estado
    repo = BRPRepository()

    if 'todouno_files' not in st.session_state:
        st.session_state.todouno_files = {'sep': None, 'pie': None, 'web': None, 'rem': None}

    # Uploader múltiple
    st.markdown("##### 📥 Cargar Archivos BRUTOS")
    st.caption("Arrastra los 3 archivos: **SEP bruto**, **PIE bruto** y **web_sostenedor**")

    uploaded_files = st.file_uploader(
        "Arrastra los 3 archivos (Excel o CSV)",
        type=['xlsx', 'xls', 'csv'],
        accept_multiple_files=True,
        key="todouno_multi_upload"
    )

    # Auto-detectar y asignar archivos
    for f in uploaded_files:
        name_lower = f.name.lower()
        if name_lower.startswith('web'):
            st.session_state.todouno_files['web'] = f
        elif name_lower.startswith('rem') or 'rem' in name_lower:
            st.session_state.todouno_files['rem'] = f
        elif name_lower.startswith('sep') or 'sep' in name_lower:
            st.session_state.todouno_files['sep'] = f
        elif name_lower.startswith('sn') or 'pie' in name_lower or 'normal' in name_lower:
            st.session_state.todouno_files['pie'] = f

    # Upload separado para REM (CSV o Excel)
    st.markdown("")
    st.markdown("##### 📋 Archivo REM (opcional)")
    st.caption("Archivo de remuneraciones para calcular horas disponibles por persona")
    rem_upload = st.file_uploader(
        "Archivo REM (CSV o Excel)",
        type=['csv', 'xlsx', 'xls'],
        key="rem_file_upload"
    )
    if rem_upload:
        st.session_state.todouno_files['rem'] = rem_upload

    # Mostrar estado de archivos detectados
    st.markdown("##### 📋 Archivos detectados")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown("**📄 SEP (bruto)**")
        f_sep = st.session_state.todouno_files['sep']
        if f_sep:
            st.success(f"✓ {f_sep.name}")
        else:
            st.warning("⬜ No detectado")

    with col2:
        st.markdown("**📄 PIE/Normal (bruto)**")
        f_pie = st.session_state.todouno_files['pie']
        if f_pie:
            st.success(f"✓ {f_pie.name}")
        else:
            st.warning("⬜ No detectado")

    with col3:
        st.markdown("**📋 MINEDUC (web)**")
        f_web = st.session_state.todouno_files['web']
        if f_web:
            st.success(f"✓ {f_web.name}")
        else:
            st.warning("⬜ No detectado")

    with col4:
        st.markdown("**📋 REM (opcional)**")
        f_rem = st.session_state.todouno_files.get('rem')
        if f_rem:
            st.success(f"✓ {f_rem.name}")
        else:
            st.info("⬜ Sin archivo REM")

    # Botón para limpiar
    if st.button("🔄 Limpiar archivos", key="btn_clear_todouno"):
        st.session_state.todouno_files = {'sep': None, 'pie': None, 'web': None, 'rem': None}
        st.rerun()

    # Verificar que están todos los archivos
    if not all([f_sep, f_pie, f_web]):
        warning_box("Carga los **3 archivos** para continuar")
        return

    st.markdown("---")

    # Configuración del procesamiento
    st.markdown("##### ⚙️ Configuración")

    col1, col2 = st.columns(2)

    with col1:
        # Selector de mes
        mes_default = datetime.now().strftime("%Y-%m")
        mes = st.text_input(
            "📅 Mes de procesamiento",
            value=mes_default,
            help="Formato: YYYY-MM (ej: 2024-01)"
        )

    with col2:
        # Checkbox para comparar
        meses_disponibles = repo.obtener_meses_disponibles()
        comparar = st.checkbox(
            "📊 Comparar con mes anterior",
            value=False,
            disabled=len(meses_disponibles) == 0,
            help="Compara con un procesamiento guardado anteriormente"
        )

        mes_anterior = None
        if comparar and meses_disponibles:
            mes_anterior = st.selectbox(
                "Selecciona mes anterior",
                options=meses_disponibles,
                key="mes_anterior_select"
            )

    # Opción para guardar en BD
    guardar_bd = st.checkbox(
        "💾 Guardar en base de datos (para comparaciones futuras)",
        value=True
    )

    st.markdown("---")

    # Botón de procesar
    if st.button("🚀 PROCESAR TODO", key="btn_todouno", width='stretch'):
        processor = IntegradoProcessor()

        progress = st.progress(0)
        status = st.empty()

        def callback(val, msg):
            progress.progress(val / 100)
            status.markdown(f"**⏳ {msg}**")

        try:
            # Crear archivos temporales
            paths = {}
            for key, f in [('sep', f_sep), ('pie', f_pie), ('web', f_web)]:
                tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                tmp.write(f.getvalue())
                paths[key] = Path(tmp.name)
                tmp.close()

            out_tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            out_path = Path(out_tmp.name)
            out_tmp.close()

            # Extraer mes para filtrar web sostenedor si tiene múltiples meses
            todouno_month_filter = None
            if mes and '-' in mes:
                todouno_month_filter = mes.split('-')[1]

            # Procesar todo
            df_result, audit = processor.process_all(
                sep_bruto_path=paths['sep'],
                pie_bruto_path=paths['pie'],
                web_sostenedor_path=paths['web'],
                output_path=out_path,
                progress_callback=callback,
                month_filter=todouno_month_filter,
            )

            progress.progress(100)
            status.markdown("**✅ ¡Procesamiento completado!**")
            st.toast("Procesamiento integrado completado", icon="🎯")

            # Mostrar alertas de columnas
            show_column_alerts(processor.brp_processor.get_column_alerts())

            # Guardar en BD si está marcado
            if guardar_bd:
                try:
                    repo.guardar_procesamiento(mes, df_result)
                    st.success(f"💾 Datos guardados en base de datos para {mes}")
                except Exception as e:
                    st.warning(f"⚠️ No se pudo guardar en BD: {str(e)}")

            # Realizar comparación si está habilitada
            comparacion = None
            if comparar and mes_anterior:
                try:
                    comparador = ComparadorMeses(repo)
                    comparacion = comparador.comparar(mes_anterior, mes)
                except Exception as e:
                    st.warning(f"⚠️ No se pudo realizar comparación: {str(e)}")

            # Leer bytes del Excel antes de que se limpie
            with open(out_path, 'rb') as f:
                excel_bytes = f.read()

            # Generar Word buffer
            word_bytes = None
            try:
                informe = InformeWord()
                word_buffer = informe.generar(
                    mes=mes,
                    df_resultado=df_result,
                    audit_log=audit,
                    comparacion=comparacion
                )
                word_bytes = word_buffer.getvalue()
            except Exception:
                pass

            # Cachear resultados en session_state
            brp_sep = df_result['BRP_SEP'].sum() if 'BRP_SEP' in df_result.columns else 0
            brp_pie = df_result['BRP_PIE'].sum() if 'BRP_PIE' in df_result.columns else 0
            brp_normal = df_result['BRP_NORMAL'].sum() if 'BRP_NORMAL' in df_result.columns else 0
            brp_total = brp_sep + brp_pie + brp_normal

            cols = df_result.columns
            daem_total_sidebar = sum(
                df_result[f'TOTAL_DAEM_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL']
                if f'TOTAL_DAEM_{s}' in cols
            )
            cpeip_total_sidebar = sum(
                df_result[f'TOTAL_CPEIP_{s}'].sum() for s in ['SEP', 'PIE', 'NORMAL']
                if f'TOTAL_CPEIP_{s}' in cols
            )

            st.session_state['last_brp_result'] = {
                'mes': mes,
                'brp_sep': brp_sep, 'brp_pie': brp_pie,
                'brp_normal': brp_normal, 'brp_total': brp_total,
                'daem_total': daem_total_sidebar, 'cpeip_total': cpeip_total_sidebar,
            }

            st.session_state['todouno_result'] = {
                'df_result': df_result,
                'excel_bytes': excel_bytes,
                'word_bytes': word_bytes,
                'audit': audit,
                'comparacion': comparacion,
                'mes': mes,
                'brp_sep': brp_sep, 'brp_pie': brp_pie,
                'brp_normal': brp_normal, 'brp_total': brp_total,
            }

            # Procesar REM si se proporcionó
            f_rem = st.session_state.todouno_files.get('rem')
            if f_rem:
                try:
                    suffix = '.csv' if f_rem.name.lower().endswith('.csv') else '.xlsx'
                    tmp_rem = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
                    tmp_rem.write(f_rem.getvalue())
                    rem_path = Path(tmp_rem.name)
                    tmp_rem.close()

                    rem_processor = REMProcessor()
                    df_rem_resumen, df_rem_detalle, rem_alertas = rem_processor.process(rem_path)

                    st.session_state['todouno_rem'] = {
                        'df_resumen': df_rem_resumen,
                        'alertas': rem_alertas,
                    }
                    st.session_state['last_rem_result'] = {
                        'personas': len(df_rem_resumen),
                        'total_horas': int(df_rem_resumen['TOTAL'].sum()),
                        'exceden': int(df_rem_resumen['EXCEDE'].sum()),
                    }
                except Exception as e:
                    st.warning(f"Error procesando REM: {format_user_error(e)}")

        except Exception as e:
            st.error(f"❌ **Error:** {format_user_error(e)}")
            import traceback
            with st.expander("Ver detalles técnicos"):
                st.code(traceback.format_exc())

    # Mostrar resultados cacheados (persisten entre reruns)
    if 'todouno_result' in st.session_state:
        cached = st.session_state['todouno_result']
        df_result = cached['df_result']
        excel_bytes = cached['excel_bytes']
        word_bytes = cached['word_bytes']
        audit = cached['audit']
        comparacion = cached['comparacion']
        cached_mes = cached['mes']
        brp_sep = cached['brp_sep']
        brp_pie = cached['brp_pie']
        brp_normal = cached['brp_normal']
        brp_total = cached['brp_total']

        st.markdown("---")

        # Métricas principales siempre visibles
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("SEP", fmt_clp(brp_sep))
            if brp_total > 0:
                st.badge(f"{100*brp_sep/brp_total:.1f}%", color="blue")
        with c2:
            st.metric("PIE", fmt_clp(brp_pie))
            if brp_total > 0:
                st.badge(f"{100*brp_pie/brp_total:.1f}%", color="green")
        with c3:
            st.metric("NORMAL", fmt_clp(brp_normal))
            if brp_total > 0:
                st.badge(f"{100*brp_normal/brp_total:.1f}%", color="orange")
        with c4:
            st.metric("TOTAL", fmt_clp(brp_total))
            st.badge("BRP Total", color="violet")

        res_tabs = st.tabs(["📈 Resumen", "📊 Gráficos", "🔍 Explorador", "📥 Descarga"])

        with res_tabs[0]:
            if brp_total > 0:
                show_desglose_daem_cpeip(df_result)

            if comparacion:
                st.markdown("---")
                st.markdown("##### 📊 Comparación con Mes Anterior")

                resumen = comparacion.get('resumen', {})

                col1, col2, col3 = st.columns(3)
                with col1:
                    diff_doc = resumen.get('docentes_actual', 0) - resumen.get('docentes_anterior', 0)
                    st.metric(
                        "Docentes",
                        f"{resumen.get('docentes_actual', 0):,}".replace(',', '.'),
                        delta=f"{diff_doc:+d}"
                    )
                with col2:
                    st.metric(
                        "Nuevos",
                        f"{resumen.get('docentes_nuevos', 0):,}".replace(',', '.')
                    )
                with col3:
                    st.metric(
                        "Salieron",
                        f"{resumen.get('docentes_salieron', 0):,}".replace(',', '.')
                    )

                col1, col2 = st.columns(2)
                with col1:
                    st.metric(
                        "BRP Total",
                        fmt_clp(resumen.get('brp_actual', 0)),
                        delta=f"{resumen.get('cambio_brp_pct', 0):+.1f}%"
                    )
                with col2:
                    cambios_monto = resumen.get('cambios_monto_significativo', 0)
                    if cambios_monto > 0:
                        st.warning(f"⚠️ {cambios_monto} docentes con cambio de monto >10%")

                with st.expander("👀 Ver detalles de comparación"):
                    nuevos = comparacion.get('docentes_nuevos', [])
                    if nuevos:
                        st.markdown("**Docentes Nuevos:**")
                        df_nuevos = pd.DataFrame(nuevos)
                        st.dataframe(df_nuevos.style.format(format_money_cols(df_nuevos)), width='stretch', hide_index=True)
                        add_table_downloads(df_nuevos, 'docentes_nuevos', 'cmp_nuevos')

                    salieron = comparacion.get('docentes_salieron', [])
                    if salieron:
                        st.markdown("**Docentes que Salieron:**")
                        df_salieron = pd.DataFrame(salieron)
                        st.dataframe(df_salieron.style.format(format_money_cols(df_salieron)), width='stretch', hide_index=True)
                        add_table_downloads(df_salieron, 'docentes_salieron', 'cmp_salieron')

                    cambios = comparacion.get('cambios_montos', [])
                    if cambios:
                        st.markdown("**Cambios Significativos de Monto:**")
                        df_cambios = pd.DataFrame(cambios)
                        st.dataframe(df_cambios.style.format(format_money_cols(df_cambios)), width='stretch', hide_index=True)
                        add_table_downloads(df_cambios, 'cambios_monto', 'cmp_cambios')

            show_multi_establishment(excel_bytes)

            show_audit_log_detailed(audit)

            # REM results
            if 'todouno_rem' in st.session_state:
                st.markdown("---")
                rem_data = st.session_state['todouno_rem']
                show_rem_results(rem_data['df_resumen'], rem_data['alertas'])

        with res_tabs[1]:
            if brp_total > 0:
                import plotly.express as px

                st.markdown("##### 📊 Distribución Visual")

                chart_type = st.selectbox(
                    "Tipo de gráfico",
                    ["Subvención (SEP/PIE/Normal)", "Auditoría"],
                    key="todo_chart_type"
                )

                if chart_type == "Subvención (SEP/PIE/Normal)":
                    df_pie_chart = pd.DataFrame({
                        'Tipo': ['SEP', 'PIE', 'NORMAL'],
                        'Monto': [brp_sep, brp_pie, brp_normal]
                    })

                    fig = px.pie(df_pie_chart, values='Monto', names='Tipo',
                                 title='Distribución por Subvención',
                                 color_discrete_sequence=['#3b82f6', '#10b981', '#f59e0b'])
                    fig.update_traces(textposition='inside', textinfo='percent+label')
                    fig.update_layout(height=400, showlegend=True, separators=',.')
                    st.plotly_chart(fig, width='stretch')
                else:
                    audit_summary = audit.get_summary()
                    n_err = audit_summary.get('errores', 0)
                    n_warn = audit_summary.get('advertencias', 0)
                    n_total_audit = audit_summary.get('total', 0)
                    cols_a = st.columns(3)
                    cols_a[0].metric("Eventos", n_total_audit)
                    cols_a[1].metric("Advertencias", n_warn)
                    cols_a[2].metric("Errores", n_err)

                    docentes_eib = len(audit.get_docentes_eib())
                    if docentes_eib > 0:
                        st.warning(f"{docentes_eib} docentes con BRP $0 (posibles EIB)")
            else:
                st.info("No hay datos para graficar.")

        with res_tabs[2]:
            show_data_explorer(df_result, key_prefix="todo")

        with res_tabs[3]:
            st.markdown("##### 📥 Descargas")

            styled_brp = _style_excel_workbook(excel_bytes)

            col_dl1, col_dl2, col_dl3 = st.columns(3)

            with col_dl1:
                st.download_button(
                    "📥 Excel Formateado",
                    data=styled_brp,
                    file_name=f"brp_distribuido_{cached_mes}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_todo_xlsx",
                )

            with col_dl2:
                if word_bytes:
                    st.download_button(
                        "📄 Informe Word",
                        data=word_bytes,
                        file_name=f"informe_brp_{cached_mes}.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        key="dl_todo_word",
                    )

            with col_dl3:
                csv_data = df_result.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    "📥 CSV Detalle",
                    data=csv_data,
                    file_name=f"brp_detalle_{cached_mes}.csv",
                    mime="text/csv",
                    key="dl_todo_csv",
                )

            with st.expander("👀 Vista previa del resultado"):
                st.dataframe(
                    df_result.head(20).style.format(format_money_cols(df_result.head(20))),
                    width='stretch', hide_index=True,
                )


@st.fragment
def _render_docentes_tab(excel_bytes):
    """Renderiza la pestaña Docentes del lote anual.

    Lee DETALLE_BRP del Excel y muestra buscador, tabla resumen por docente,
    detalle mensual, gráfico de tendencia y detección multi-escuela.
    """
    import plotly.graph_objects as go

    df = _read_detalle_brp(excel_bytes)
    if df.empty:
        st.info("No hay datos de detalle BRP disponibles.")
        return

    # Detectar columnas usando búsqueda parcial (como show_data_explorer)
    col_rut = _find_col(df, 'rut')
    col_nombre = _find_col(df, 'nombre_completo', 'nombre')
    col_rbd = _find_col(df, 'rbd')
    has_mes = 'MES' in df.columns

    if not col_rut:
        st.warning("No se encontró columna de RUT en el detalle BRP.")
        return

    # Formatear RUT
    df['RUT_FMT'] = df[col_rut].apply(format_rut)

    # Columnas BRP
    brp_cols = [c for c in ['BRP_SEP', 'BRP_PIE', 'BRP_NORMAL', 'BRP_TOTAL'] if c in df.columns]
    daem_cols = [c for c in df.columns if c.startswith('DAEM_') and c != 'DAEM_TOTAL']
    cpeip_cols = [c for c in df.columns if c.startswith('CPEIP_') and c != 'CPEIP_TOTAL']

    # --- Buscador ---
    st.markdown("### Buscar Docente")
    col_search, col_mes_filter = st.columns([2, 1])
    with col_search:
        search = st.text_input("Buscar por nombre, apellido o RUT", key="doc_search", placeholder="Ej: 12345678-9 o García")
    with col_mes_filter:
        if has_mes:
            meses_disp = sorted(df['MES'].dropna().unique().tolist())
            mes_filter = st.multiselect("Filtrar por mes", meses_disp, default=meses_disp, key="doc_mes_filter")
        else:
            mes_filter = None

    # Aplicar filtros
    df_filtered = df.copy()
    if has_mes and mes_filter is not None:
        df_filtered = df_filtered[df_filtered['MES'].isin(mes_filter)]
    if search:
        search_lower = search.lower()
        mask = df_filtered['RUT_FMT'].str.lower().str.contains(search_lower, na=False)
        if col_nombre:
            mask = mask | df_filtered[col_nombre].astype(str).str.lower().str.contains(search_lower, na=False)
        df_filtered = df_filtered[mask]

    if df_filtered.empty:
        st.warning("No se encontraron docentes con los filtros aplicados.")
        return

    # --- Tabla resumen agrupada por RUT ---
    st.markdown("### Resumen por Docente")
    agg_dict = {}
    if col_nombre:
        agg_dict[col_nombre] = (col_nombre, 'first')
    for c in brp_cols:
        agg_dict[c] = (c, 'sum')
    if has_mes:
        agg_dict['MESES_PRESENTES'] = ('MES', 'nunique')
    if col_rbd:
        agg_dict['NUM_ESCUELAS'] = (col_rbd, 'nunique')

    df_resumen = df_filtered.groupby(col_rut).agg(**agg_dict).reset_index()
    df_resumen['RUT_FMT'] = df_resumen[col_rut].apply(format_rut)

    # Reordenar columnas
    display_cols = ['RUT_FMT']
    if col_nombre and col_nombre in df_resumen.columns:
        display_cols.append(col_nombre)
    display_cols += [c for c in brp_cols if c in df_resumen.columns]
    if 'MESES_PRESENTES' in df_resumen.columns:
        display_cols.append('MESES_PRESENTES')
    if 'NUM_ESCUELAS' in df_resumen.columns:
        display_cols.append('NUM_ESCUELAS')

    df_resumen_display = df_resumen[display_cols].rename(columns={'RUT_FMT': 'RUT'}).copy()
    df_resumen_display = append_totals_row(df_resumen_display, label_col='RUT', label='TOTAL',
                                           skip_cols={col_nombre} if col_nombre else None)

    # Badge multi-escuela
    n_multi = len(df_resumen[df_resumen.get('NUM_ESCUELAS', pd.Series([0])) > 1]) if 'NUM_ESCUELAS' in df_resumen.columns else 0
    if n_multi > 0:
        st.info(f"**{n_multi}** docente(s) trabajan en 2+ establecimientos")

    render_generic_interactive_table(df_resumen_display, key="doc_resumen")
    st.caption(f"{len(df_resumen)} docentes encontrados")

    # --- Selector de docente ---
    st.markdown("---")
    st.markdown("### Detalle Individual")

    rut_options = df_resumen[col_rut].tolist()
    rut_labels = []
    for _, r in df_resumen.iterrows():
        nombre = r.get(col_nombre, '') if col_nombre else ''
        n_esc = int(r.get('NUM_ESCUELAS', 1)) if 'NUM_ESCUELAS' in df_resumen.columns else 1
        label = f"{r['RUT_FMT']} — {nombre}"
        if n_esc > 1:
            label += f" ({n_esc} escuelas)"
        rut_labels.append(label)

    sel_idx = st.selectbox(
        "Seleccionar docente", range(len(rut_labels)),
        format_func=lambda i: rut_labels[i], key="doc_sel"
    )
    sel_rut = rut_options[sel_idx]
    df_docente = df_filtered[df_filtered[col_rut] == sel_rut].copy()

    # --- Métricas ---
    mc = st.columns(min(len(brp_cols) + 2, 6))
    for i, col in enumerate(brp_cols):
        with mc[i]:
            st.metric(col.replace('BRP_', 'BRP '), fmt_clp(df_docente[col].sum()))
    idx = len(brp_cols)
    if daem_cols:
        daem_total = sum(df_docente[c].sum() for c in daem_cols if c in df_docente.columns)
        with mc[min(idx, len(mc) - 1)]:
            st.metric("DAEM Total", fmt_clp(daem_total))
        idx += 1
    if cpeip_cols:
        cpeip_total = sum(df_docente[c].sum() for c in cpeip_cols if c in df_docente.columns)
        with mc[min(idx, len(mc) - 1)]:
            st.metric("CPEIP Total", fmt_clp(cpeip_total))

    # --- Tabla mensual ---
    st.markdown("##### Detalle mensual")
    monthly_cols = []
    if has_mes:
        monthly_cols.append('MES')
    if col_rbd:
        monthly_cols.append(col_rbd)
    # Agregar nombre de escuela si disponible
    escuelas = get_rbd_map()
    if col_rbd and escuelas:
        df_docente['ESCUELA'] = df_docente[col_rbd].apply(
            lambda x: escuelas.get(str(x).split('.')[0].split('-')[0].strip(), '') if pd.notna(x) else ''
        )
        monthly_cols.append('ESCUELA')
    col_tipo = next((c for c in df_docente.columns if 'TIPO' in c.upper() and 'PAGO' in c.upper()), None)
    if col_tipo:
        monthly_cols.append(col_tipo)
    hr_cols = [c for c in ['HORAS_SEP', 'HORAS_PIE', 'HORAS_SN'] if c in df_docente.columns]
    monthly_cols += hr_cols
    monthly_cols += brp_cols
    monthly_cols = [c for c in monthly_cols if c in df_docente.columns]

    df_monthly = df_docente[monthly_cols].copy()
    if has_mes:
        df_monthly = append_totals_row(df_monthly, label_col='MES', label='TOTAL',
                                       skip_cols={'ESCUELA', col_tipo} if col_tipo else {'ESCUELA'})
    render_generic_interactive_table(df_monthly, key="doc_monthly")

    # Botón PDF para tabla mensual del docente
    try:
        sel_nombre = df_resumen.iloc[sel_idx].get(col_nombre, '') if col_nombre else ''
        pdf_doc_bytes = generate_pdf_from_df(
            df_monthly,
            title=f"Detalle Docente - {sel_nombre}",
            orientation='L',
        )
        st.download_button(
            "📥 Descargar PDF del docente",
            data=pdf_doc_bytes,
            file_name=f"docente_{sel_rut}.pdf",
            mime="application/pdf",
            key="dl_doc_pdf",
        )
    except Exception:
        pass

    # --- Gráfico de tendencia ---
    if has_mes and len(df_docente) > 1:
        st.markdown("##### Tendencia mensual")
        df_chart = df_docente.copy()
        if 'MES_NUM' in df_chart.columns:
            df_chart = df_chart.sort_values('MES_NUM')

        # Agrupar por mes (puede haber múltiples registros por mes si multi-escuela)
        chart_agg = {}
        for c in brp_cols:
            chart_agg[c] = 'sum'
        df_trend = df_chart.groupby('MES').agg(chart_agg).reset_index()

        # Ordenar por MES_NUM si disponible
        if 'MES_NUM' in df_chart.columns:
            mes_order = df_chart.drop_duplicates('MES')[['MES', 'MES_NUM']].sort_values('MES_NUM')
            df_trend = df_trend.set_index('MES').reindex(mes_order['MES'].values).reset_index()

        fig = go.Figure()
        if 'BRP_TOTAL' in df_trend.columns:
            fig.add_trace(go.Scatter(
                x=df_trend['MES'], y=df_trend['BRP_TOTAL'],
                mode='lines+markers', name='BRP Total',
                line=dict(color='#1e293b', width=3),
            ))
        colors = {'BRP_SEP': '#3b82f6', 'BRP_PIE': '#10b981', 'BRP_NORMAL': '#f59e0b'}
        for col in ['BRP_SEP', 'BRP_PIE', 'BRP_NORMAL']:
            if col in df_trend.columns:
                fig.add_trace(go.Bar(
                    x=df_trend['MES'], y=df_trend[col],
                    name=col.replace('BRP_', ''), marker_color=colors.get(col, '#94a3b8'),
                ))
        fig.update_layout(
            barmode='stack', height=350,
            yaxis_title='Monto ($)',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
            separators=',.',
        )
        st.plotly_chart(fig, width='stretch')

    # --- Multi-escuela ---
    if col_rbd and 'NUM_ESCUELAS' in df_resumen.columns:
        sel_info = df_resumen[df_resumen[col_rut] == sel_rut]
        if not sel_info.empty and int(sel_info.iloc[0].get('NUM_ESCUELAS', 1)) > 1:
            st.markdown("##### Distribución por Establecimiento")
            agg_esc = {}
            for c in brp_cols:
                agg_esc[c] = (c, 'sum')
            if has_mes:
                agg_esc['MESES'] = ('MES', 'nunique')
            df_by_esc = df_docente.groupby(col_rbd).agg(**agg_esc).reset_index()
            if escuelas:
                df_by_esc['ESCUELA'] = df_by_esc[col_rbd].apply(
                    lambda x: escuelas.get(str(x).split('.')[0].split('-')[0].strip(), '')
                )
            df_by_esc = append_totals_row(df_by_esc, label_col=col_rbd, label='TOTAL',
                                          skip_cols={'ESCUELA'})
            render_generic_interactive_table(df_by_esc, key="doc_multi_esc")


def _style_excel_workbook(excel_bytes):
    """Aplica formato profesional a un Excel existente (headers, CLP, totales).

    Returns:
        bytes del Excel estilizado.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = load_workbook(BytesIO(excel_bytes))
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    clp_format = '#,##0'

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue

        # Detectar columnas monetarias (numéricas, no horas/conteos)
        money_cols = set()
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            col_name = str(cell.value or '')
            headers[col_idx] = col_name
            # Header styling
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

            if col_name not in NON_MONEY_COLS and col_name not in ('RBD', 'MES', 'MES_NUM', 'MOTIVO', 'DETALLE', 'ACCION', 'RUT', 'NOMBRE', 'APELLIDOS', 'NOMBRE_COMPLETO', 'ESCUELA', 'MULTI_ESTABLECIMIENTO', 'TIPO_FILA', 'TRAMO'):
                # Verificar si la col tiene números
                sample_cell = ws.cell(row=2, column=col_idx)
                if isinstance(sample_cell.value, (int, float)):
                    money_cols.add(col_idx)

        # Data rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx in money_cols:
                    cell.number_format = clp_format
                    cell.alignment = Alignment(horizontal='right')

        # Auto-ancho
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_len = len(str(headers.get(col_idx, '')))
            for row_idx in range(2, min(ws.max_row + 1, 52)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    val_len = len(str(cell.value))
                    if col_idx in money_cols and isinstance(cell.value, (int, float)):
                        val_len = len(f'{int(cell.value):,}') + 1
                    max_len = max(max_len, val_len)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 30)

        # Freeze header
        ws.freeze_panes = ws.cell(row=2, column=1)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _display_lote_anual_results(stats, excel_data, anio):
    """Muestra resultados detallados del procesamiento anual.

    Args:
        stats: dict con summaries y metadata del procesamiento.
        excel_data: bytes del Excel generado o Path al archivo.
        anio: año del procesamiento.
    """
    import plotly.graph_objects as go
    import plotly.express as px

    # Soporte para bytes o Path
    if isinstance(excel_data, bytes):
        excel_src = BytesIO(excel_data)
        excel_bytes = excel_data
    else:
        excel_src = str(excel_data)
        with open(excel_data, 'rb') as f:
            excel_bytes = f.read()

    summaries = [s for s in stats['summaries'] if 'ERROR' not in s]
    errores = [s for s in stats['summaries'] if 'ERROR' in s]

    if not summaries:
        st.warning("No se procesó ningún mes exitosamente.")
        return

    success_box(
        f"Lote anual completado: **{stats['meses_procesados']}** meses procesados"
    )

    # --- Totales anuales desde summaries ---
    brp_sep = sum(s.get('BRP_SEP', 0) for s in summaries)
    brp_pie = sum(s.get('BRP_PIE', 0) for s in summaries)
    brp_normal = sum(s.get('BRP_NORMAL', 0) for s in summaries)
    brp_total = brp_sep + brp_pie + brp_normal

    daem_sep = sum(s.get('DAEM_SEP', 0) for s in summaries)
    daem_pie = sum(s.get('DAEM_PIE', 0) for s in summaries)
    daem_normal = sum(s.get('DAEM_NORMAL', 0) for s in summaries)
    daem_total = daem_sep + daem_pie + daem_normal

    cpeip_sep = sum(s.get('CPEIP_SEP', 0) for s in summaries)
    cpeip_pie = sum(s.get('CPEIP_PIE', 0) for s in summaries)
    cpeip_normal = sum(s.get('CPEIP_NORMAL', 0) for s in summaries)
    cpeip_total = cpeip_sep + cpeip_pie + cpeip_normal

    recon_sep = sum(s.get('RECON_SEP', 0) for s in summaries)
    recon_pie = sum(s.get('RECON_PIE', 0) for s in summaries)
    recon_normal = sum(s.get('RECON_NORMAL', 0) for s in summaries)

    tramo_sep = sum(s.get('TRAMO_SEP', 0) for s in summaries)
    tramo_pie = sum(s.get('TRAMO_PIE', 0) for s in summaries)
    tramo_normal = sum(s.get('TRAMO_NORMAL', 0) for s in summaries)

    prior_sep = sum(s.get('PRIOR_SEP', 0) for s in summaries)
    prior_pie = sum(s.get('PRIOR_PIE', 0) for s in summaries)
    prior_normal = sum(s.get('PRIOR_NORMAL', 0) for s in summaries)

    table_rows = []
    for s in sorted(summaries, key=lambda x: x.get('MES_NUM', '00')):
        table_rows.append({
            'MES': s['MES'],
            'DOCENTES': s.get('DOCENTES_BRP', 0),
            'BRP_SEP': int(s.get('BRP_SEP', 0)),
            'BRP_PIE': int(s.get('BRP_PIE', 0)),
            'BRP_NORMAL': int(s.get('BRP_NORMAL', 0)),
            'BRP_TOTAL': int(s.get('BRP_TOTAL', 0)),
            'DAEM_TOTAL': int(s.get('DAEM_TOTAL', 0)),
            'CPEIP_TOTAL': int(s.get('CPEIP_TOTAL', 0)),
        })

    st.markdown("---")

    # Métricas principales siempre visibles (fuera de tabs)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("SEP", fmt_clp(brp_sep))
        if brp_total > 0:
            st.badge(f"{100*brp_sep/brp_total:.1f}%", color="blue")
    with c2:
        st.metric("PIE", fmt_clp(brp_pie))
        if brp_total > 0:
            st.badge(f"{100*brp_pie/brp_total:.1f}%", color="green")
    with c3:
        st.metric("NORMAL", fmt_clp(brp_normal))
        if brp_total > 0:
            st.badge(f"{100*brp_normal/brp_total:.1f}%", color="orange")
    with c4:
        st.metric("TOTAL", fmt_clp(brp_total))
        st.badge("BRP Total", color="violet")

    res_tabs = st.tabs(["📈 Resumen", "🏫 Por Establecimiento", "👥 Docentes", "📋 Horas", "📊 Gráficos", "🔍 Explorador", "⚠️ Alertas", "📥 Descarga"])

    with res_tabs[0]:
        st.markdown("### Resumen Anual")

        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Paga DAEM", fmt_clp(daem_total))
        with c2:
            st.metric("Paga CPEIP", fmt_clp(cpeip_total))
        with c3:
            total_docentes = max(s.get('DOCENTES_BRP', 0) for s in summaries)
            total_estab = max(s.get('ESTABLECIMIENTOS', 0) for s in summaries)
            st.metric("Docentes", total_docentes)
            st.caption(f"{total_estab} establecimientos")

        # Tabla por mes
        st.markdown("---")
        st.markdown("### Detalle por Mes")

        df_table = pd.DataFrame(table_rows)
        totals = {'MES': 'TOTAL'}
        for col in df_table.columns:
            if col != 'MES':
                totals[col] = df_table[col].sum()
        df_table = pd.concat([df_table, pd.DataFrame([totals])], ignore_index=True)

        money_cols = [c for c in df_table.columns if c not in ('MES', 'DOCENTES')]
        st.dataframe(
            df_table.style.format({col: fmt_clp_style for col in money_cols}),
            width='stretch',
            hide_index=True,
        )
        add_table_downloads(df_table, f'detalle_por_mes_{anio}', 'lote_mes')

        # Detalle Pagador por Subvención
        st.markdown("---")
        st.markdown("### Detalle Pagador por Subvención")

        detalle_pagador = pd.DataFrame([
            {'Pagador': 'DAEM', 'SEP ($)': daem_sep, 'PIE ($)': daem_pie, 'Normal ($)': daem_normal, 'Total ($)': daem_total},
            {'Pagador': 'CPEIP', 'SEP ($)': cpeip_sep, 'PIE ($)': cpeip_pie, 'Normal ($)': cpeip_normal, 'Total ($)': cpeip_total},
            {'Pagador': 'TOTAL', 'SEP ($)': brp_sep, 'PIE ($)': brp_pie, 'Normal ($)': brp_normal, 'Total ($)': brp_total},
        ])
        st.dataframe(
            detalle_pagador.style.format({
                'SEP ($)': fmt_clp_style, 'PIE ($)': fmt_clp_style,
                'Normal ($)': fmt_clp_style, 'Total ($)': fmt_clp_style
            }),
            width='stretch', hide_index=True,
        )
        add_table_downloads(detalle_pagador, f'detalle_pagador_{anio}', 'lote_pagador')

        with st.expander("Desglose por Concepto (Reconocimiento, Tramo, Prioritarios)"):
            df_concepto = pd.DataFrame([
                {'Concepto': 'Reconocimiento', 'SEP ($)': recon_sep, 'PIE ($)': recon_pie, 'Normal ($)': recon_normal, 'Total ($)': recon_sep + recon_pie + recon_normal},
                {'Concepto': 'Tramo', 'SEP ($)': tramo_sep, 'PIE ($)': tramo_pie, 'Normal ($)': tramo_normal, 'Total ($)': tramo_sep + tramo_pie + tramo_normal},
                {'Concepto': 'Prioritarios', 'SEP ($)': prior_sep, 'PIE ($)': prior_pie, 'Normal ($)': prior_normal, 'Total ($)': prior_sep + prior_pie + prior_normal},
                {'Concepto': 'TOTAL',
                 'SEP ($)': recon_sep + tramo_sep + prior_sep,
                 'PIE ($)': recon_pie + tramo_pie + prior_pie,
                 'Normal ($)': recon_normal + tramo_normal + prior_normal,
                 'Total ($)': recon_sep + recon_pie + recon_normal + tramo_sep + tramo_pie + tramo_normal + prior_sep + prior_pie + prior_normal},
            ])
            st.dataframe(
                df_concepto.style.format({
                    'SEP ($)': fmt_clp_style, 'PIE ($)': fmt_clp_style,
                    'Normal ($)': fmt_clp_style, 'Total ($)': fmt_clp_style
                }),
                width='stretch', hide_index=True,
            )
            add_table_downloads(df_concepto, f'desglose_concepto_{anio}', 'lote_concepto')
            st.caption("**Reconocimiento** = Reconocimiento Profesional (DAEM + CPEIP)")
            st.caption("**Tramo** = Tramo de Desarrollo Profesional (DAEM + CPEIP)")
            st.caption("**Prioritarios** = Asignación Alumnos Prioritarios (solo CPEIP)")

        # Detalle por mes
        st.markdown("---")
        st.markdown("### Detalle por Mes (Todo en Uno)")

        for s in sorted(summaries, key=lambda x: x.get('MES_NUM', '00')):
            mes_name = s['MES']
            mes_brp = int(s.get('BRP_TOTAL', 0))
            with st.expander(f"{mes_name} — BRP Total: {fmt_clp(mes_brp)}"):
                mc1, mc2, mc3, mc4 = st.columns(4)
                with mc1:
                    st.metric("BRP SEP", fmt_clp(int(s.get('BRP_SEP', 0))))
                with mc2:
                    st.metric("BRP PIE", fmt_clp(int(s.get('BRP_PIE', 0))))
                with mc3:
                    st.metric("BRP Normal", fmt_clp(int(s.get('BRP_NORMAL', 0))))
                with mc4:
                    st.metric("Docentes", s.get('DOCENTES_BRP', 0))

                mc5, mc6, mc7 = st.columns(3)
                with mc5:
                    st.metric("DAEM", fmt_clp(int(s.get('DAEM_TOTAL', 0))))
                with mc6:
                    st.metric("CPEIP", fmt_clp(int(s.get('CPEIP_TOTAL', 0))))
                with mc7:
                    st.metric("Establecimientos", s.get('ESTABLECIMIENTOS', 0))

                _r_sep = int(s.get('RECON_SEP', 0)); _r_pie = int(s.get('RECON_PIE', 0)); _r_nor = int(s.get('RECON_NORMAL', 0))
                _t_sep = int(s.get('TRAMO_SEP', 0)); _t_pie = int(s.get('TRAMO_PIE', 0)); _t_nor = int(s.get('TRAMO_NORMAL', 0))
                _p_sep = int(s.get('PRIOR_SEP', 0)); _p_pie = int(s.get('PRIOR_PIE', 0)); _p_nor = int(s.get('PRIOR_NORMAL', 0))
                df_mes_concepto = pd.DataFrame([
                    {'Concepto': 'Reconocimiento', 'SEP ($)': _r_sep, 'PIE ($)': _r_pie, 'Normal ($)': _r_nor, 'Total ($)': _r_sep + _r_pie + _r_nor},
                    {'Concepto': 'Tramo', 'SEP ($)': _t_sep, 'PIE ($)': _t_pie, 'Normal ($)': _t_nor, 'Total ($)': _t_sep + _t_pie + _t_nor},
                    {'Concepto': 'Prioritarios', 'SEP ($)': _p_sep, 'PIE ($)': _p_pie, 'Normal ($)': _p_nor, 'Total ($)': _p_sep + _p_pie + _p_nor},
                    {'Concepto': 'TOTAL',
                     'SEP ($)': _r_sep + _t_sep + _p_sep, 'PIE ($)': _r_pie + _t_pie + _p_pie,
                     'Normal ($)': _r_nor + _t_nor + _p_nor, 'Total ($)': _r_sep + _r_pie + _r_nor + _t_sep + _t_pie + _t_nor + _p_sep + _p_pie + _p_nor},
                ])
                st.dataframe(
                    df_mes_concepto.style.format({
                        'SEP ($)': fmt_clp_style, 'PIE ($)': fmt_clp_style,
                        'Normal ($)': fmt_clp_style, 'Total ($)': fmt_clp_style
                    }),
                    width='stretch', hide_index=True,
                )

                _d_sep = int(s.get('DAEM_SEP', 0)); _d_pie = int(s.get('DAEM_PIE', 0)); _d_nor = int(s.get('DAEM_NORMAL', 0)); _d_tot = int(s.get('DAEM_TOTAL', 0))
                _c_sep = int(s.get('CPEIP_SEP', 0)); _c_pie = int(s.get('CPEIP_PIE', 0)); _c_nor = int(s.get('CPEIP_NORMAL', 0)); _c_tot = int(s.get('CPEIP_TOTAL', 0))
                df_mes_pagador = pd.DataFrame([
                    {'Pagador': 'DAEM', 'SEP ($)': _d_sep, 'PIE ($)': _d_pie, 'Normal ($)': _d_nor, 'Total ($)': _d_tot},
                    {'Pagador': 'CPEIP', 'SEP ($)': _c_sep, 'PIE ($)': _c_pie, 'Normal ($)': _c_nor, 'Total ($)': _c_tot},
                    {'Pagador': 'TOTAL', 'SEP ($)': _d_sep + _c_sep, 'PIE ($)': _d_pie + _c_pie,
                     'Normal ($)': _d_nor + _c_nor, 'Total ($)': _d_tot + _c_tot},
                ])
                st.dataframe(
                    df_mes_pagador.style.format({
                        'SEP ($)': fmt_clp_style, 'PIE ($)': fmt_clp_style,
                        'Normal ($)': fmt_clp_style, 'Total ($)': fmt_clp_style
                    }),
                    width='stretch', hide_index=True,
                )

                if s.get('COSTO_EIB', 0) > 0:
                    st.metric("Costo EIB", fmt_clp(int(s['COSTO_EIB'])))

        # Sábanas por subvención
        if stats.get('tiene_detalle_sep_pie'):
            st.markdown("---")
            st.markdown("### Detalle por Subvención")

            with st.expander("Sábana SEP (columnas _SEP)"):
                try:
                    excel_src_sep = BytesIO(excel_bytes) if isinstance(excel_data, bytes) else str(excel_data)
                    df_sep = pd.read_excel(excel_src_sep, sheet_name='DETALLE_SEP', engine='openpyxl')
                    meses = df_sep['MES'].unique().tolist()
                    mes_sel = st.selectbox("Mes", ["Todos"] + meses, key="sep_mes")
                    if mes_sel != "Todos":
                        df_sep = df_sep[df_sep['MES'] == mes_sel]
                    df_sep = append_totals_row(df_sep, label_col='MES' if 'MES' in df_sep.columns else df_sep.columns[0], label='TOTAL')
                    st.dataframe(
                        df_sep.style.format(format_money_cols(df_sep)),
                        width='stretch', hide_index=True,
                    )
                    add_table_downloads(df_sep, f'sabana_SEP_{anio}', 'lote_sep_dl')
                except Exception:
                    st.info("No hay datos de detalle SEP disponibles.")

            with st.expander("Sábana PIE + Normal (columnas PIE/SN/_nuevo)"):
                try:
                    excel_src_pie = BytesIO(excel_bytes) if isinstance(excel_data, bytes) else str(excel_data)
                    df_pie = pd.read_excel(excel_src_pie, sheet_name='DETALLE_PIE', engine='openpyxl')
                    meses = df_pie['MES'].unique().tolist()
                    mes_sel = st.selectbox("Mes", ["Todos"] + meses, key="pie_mes")
                    if mes_sel != "Todos":
                        df_pie = df_pie[df_pie['MES'] == mes_sel]
                    df_pie = append_totals_row(df_pie, label_col='MES' if 'MES' in df_pie.columns else df_pie.columns[0], label='TOTAL')
                    st.dataframe(
                        df_pie.style.format(format_money_cols(df_pie)),
                        width='stretch', hide_index=True,
                    )
                    add_table_downloads(df_pie, f'sabana_PIE_{anio}', 'lote_pie_dl')
                except Exception:
                    st.info("No hay datos de detalle PIE disponibles.")

        # Resumen por RBD
        with st.expander("Resumen por Establecimiento (RBD)"):
            try:
                excel_src_rbd = BytesIO(excel_bytes) if isinstance(excel_data, bytes) else str(excel_data)
                df_rbd = pd.read_excel(excel_src_rbd, sheet_name='POR_RBD', engine='openpyxl')
                if not df_rbd.empty:
                    df_rbd = append_totals_row(df_rbd, label_col='RBD', label='TOTAL')
                    st.dataframe(
                        df_rbd.style.format(format_money_cols(df_rbd, exclude_cols={'RBD'})),
                        width='stretch', hide_index=True,
                    )
                    add_table_downloads(df_rbd, f'resumen_por_RBD_{anio}', 'lote_rbd_dl')
            except Exception:
                st.info("No hay datos por RBD disponibles.")

    with res_tabs[1]:
        st.markdown("### CPEIP por Establecimiento y Mes")
        st.caption("Transferencias directas CPEIP agrupadas por RBD. Click en celda para seleccionar, Ctrl/Cmd+Click para multi-selección.")

        # Opción para cargar CSV externo (ej: tabla pivot exportada)
        csv_externo = st.file_uploader(
            "Cargar CSV externo de establecimientos (opcional)",
            type=['csv'], key="estab_csv_ext",
            help="Acepta CSV con formato pivot: RBD, ESTABLECIMIENTO, y 3 columnas por mes (Recon, Tramo, Prior)"
        )

        df_estab = None
        if csv_externo:
            df_estab = parse_establishment_csv(csv_externo.getvalue())
            if df_estab is not None:
                st.success(f"CSV cargado: {len(df_estab)} establecimientos")
        if df_estab is None:
            df_estab = build_establishment_table(excel_bytes, summaries)

        if df_estab is not None and len(df_estab) > 1:
            render_interactive_table(df_estab, key="estab_anual")

            # Descargas
            st.markdown("---")
            col_dl1, col_dl2, col_dl3 = st.columns(3)
            with col_dl1:
                excel_estab = to_styled_excel(df_estab, sheet_name='POR_ESTABLECIMIENTO',
                                              title=f'CPEIP por Establecimiento - {anio}')
                st.download_button(
                    "📥 Excel", data=excel_estab,
                    file_name=f"CPEIP_por_establecimiento_{anio}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_estab_xlsx",
                )
            with col_dl2:
                csv_bytes = df_estab.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    "📥 CSV", data=csv_bytes,
                    file_name=f"CPEIP_por_establecimiento_{anio}.csv",
                    mime="text/csv",
                    key="dl_estab_csv",
                )
            with col_dl3:
                try:
                    pdf_bytes = generate_pdf_from_df(
                        df_estab, title=f"CPEIP por Establecimiento - {anio}",
                    )
                    st.download_button(
                        "📥 PDF", data=pdf_bytes,
                        file_name=f"CPEIP_por_establecimiento_{anio}.pdf",
                        mime="application/pdf",
                        key="dl_estab_pdf",
                    )
                except Exception:
                    st.caption("PDF no disponible")
        else:
            st.info("No hay datos de detalle BRP para generar la tabla por establecimiento.")

    with res_tabs[2]:
        _render_docentes_tab(excel_bytes)

    with res_tabs[3]:
        _render_horas_tab(excel_bytes)

    with res_tabs[4]:
        st.markdown("### Gráficos Anuales")

        df_chart = pd.DataFrame(table_rows)

        col_g1, col_g2 = st.columns(2)

        with col_g1:
            fig1 = go.Figure()
            fig1.add_trace(go.Bar(x=df_chart['MES'], y=df_chart['BRP_SEP'], name='SEP', marker_color='#3b82f6'))
            fig1.add_trace(go.Bar(x=df_chart['MES'], y=df_chart['BRP_PIE'], name='PIE', marker_color='#10b981'))
            fig1.add_trace(go.Bar(x=df_chart['MES'], y=df_chart['BRP_NORMAL'], name='Normal', marker_color='#f59e0b'))
            fig1.update_layout(
                title='BRP por Subvención (mensual)', barmode='stack', height=400,
                yaxis_title='Monto ($)',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
                separators=',.',
            )
            st.plotly_chart(fig1, width='stretch')

        with col_g2:
            fig2 = go.Figure()
            fig2.add_trace(go.Bar(x=df_chart['MES'], y=df_chart['DAEM_TOTAL'], name='DAEM (Subvención)', marker_color='#6366f1'))
            fig2.add_trace(go.Bar(x=df_chart['MES'], y=df_chart['CPEIP_TOTAL'], name='CPEIP (Transferencia)', marker_color='#ec4899'))
            fig2.update_layout(
                title='DAEM vs CPEIP (mensual)', barmode='group', height=400,
                yaxis_title='Monto ($)',
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
                separators=',.',
            )
            st.plotly_chart(fig2, width='stretch')

        col_p1, col_p2 = st.columns(2)
        with col_p1:
            df_subv = pd.DataFrame({
                'Subvención': ['SEP', 'PIE', 'Normal'],
                'Monto': [brp_sep, brp_pie, brp_normal]
            })
            fig3 = px.pie(df_subv, values='Monto', names='Subvención',
                          title='Distribución Anual por Subvención',
                          color_discrete_sequence=['#3b82f6', '#10b981', '#f59e0b'])
            fig3.update_traces(textposition='inside', textinfo='percent+label+value',
                               texttemplate='%{label}<br>$%{value:,.0f}<br>(%{percent})')
            fig3.update_layout(height=350, showlegend=False, separators=',.')
            st.plotly_chart(fig3, width='stretch')

        with col_p2:
            df_pagador = pd.DataFrame({
                'Pagador': ['DAEM (Subvención)', 'CPEIP (Transferencia)'],
                'Monto': [daem_total, cpeip_total]
            })
            fig4 = px.pie(df_pagador, values='Monto', names='Pagador',
                          title='Distribución Anual DAEM vs CPEIP',
                          color_discrete_sequence=['#6366f1', '#ec4899'])
            fig4.update_traces(textposition='inside', textinfo='percent+label+value',
                               texttemplate='%{label}<br>$%{value:,.0f}<br>(%{percent})')
            fig4.update_layout(height=350, showlegend=False, separators=',.')
            st.plotly_chart(fig4, width='stretch')

    with res_tabs[5]:
        # Explorador interactivo
        try:
            excel_src_brp = BytesIO(excel_bytes) if isinstance(excel_data, bytes) else str(excel_data)
            df_detalle_brp = pd.read_excel(excel_src_brp, sheet_name='DETALLE_BRP', engine='openpyxl')
            if not df_detalle_brp.empty:
                show_data_explorer(df_detalle_brp, key_prefix="lote")
            else:
                st.info("No hay datos de detalle BRP disponibles.")
        except Exception:
            st.info("No hay datos de detalle BRP disponibles.")

    with res_tabs[6]:
        st.markdown("### Alertas y Revisión")

        total_eib_docentes = sum(s.get('DOCENTES_EIB', 0) for s in summaries)
        eib_total_costo = stats.get('eib_total_anual', 0)

        # Leer REVISAR
        df_revisar = pd.DataFrame()
        try:
            excel_src_rev = BytesIO(excel_bytes) if isinstance(excel_data, bytes) else str(excel_data)
            df_revisar = pd.read_excel(excel_src_rev, sheet_name='REVISAR', engine='openpyxl')
        except Exception:
            pass

        n_revisar = len(df_revisar)
        n_exceden = 0
        n_sin_liq = 0
        if not df_revisar.empty and 'MOTIVO' in df_revisar.columns:
            n_exceden = len(df_revisar[df_revisar['MOTIVO'].str.contains('44', na=False)])
            n_sin_liq = len(df_revisar[df_revisar['MOTIVO'].str.contains('SIN', na=False, case=False)])

        ac1, ac2, ac3, ac4, ac5 = st.columns(5)
        with ac1:
            if errores:
                st.metric("Meses con Error", len(errores))
            else:
                st.metric("Meses con Error", 0)
        with ac2:
            st.metric("A Revisar", n_revisar)
        with ac3:
            st.metric("Exceden 44 hrs", n_exceden)
        with ac4:
            st.metric("Sin Liquidación", n_sin_liq)
        with ac5:
            st.metric("Docentes EIB", total_eib_docentes)

        if errores:
            st.error(f"**{len(errores)} mes(es) con errores de procesamiento:**")
            for s in errores:
                st.markdown(f"- **{s['MES']}**: {s['ERROR']}")

        if eib_total_costo > 0:
            st.warning(
                f"**EIB:** {total_eib_docentes} docentes con posible BRP $0 "
                f"(costo EIB anual: {fmt_clp(eib_total_costo)})"
            )

        if not df_revisar.empty:
            st.markdown("##### Docentes a Revisar")

            if 'MOTIVO' in df_revisar.columns:
                motivos = sorted(df_revisar['MOTIVO'].dropna().unique().tolist())
                motivo_sel = st.multiselect(
                    "Filtrar por motivo", motivos, default=motivos,
                    key="lote_revisar_motivo"
                )
                df_revisar_f = df_revisar[df_revisar['MOTIVO'].isin(motivo_sel)]
            else:
                df_revisar_f = df_revisar

            if 'MES' in df_revisar_f.columns:
                meses_rev = sorted(df_revisar_f['MES'].dropna().unique().tolist())
                mes_rev_sel = st.multiselect(
                    "Filtrar por mes", meses_rev, default=meses_rev,
                    key="lote_revisar_mes"
                )
                df_revisar_f = df_revisar_f[df_revisar_f['MES'].isin(mes_rev_sel)]

            st.dataframe(
                df_revisar_f.style.format(format_money_cols(df_revisar_f)),
                width='stretch', hide_index=True,
            )
            st.caption(f"Mostrando {len(df_revisar_f)} de {n_revisar} registros")
            add_table_downloads(df_revisar_f, f'docentes_a_revisar_{anio}', 'lote_revisar_dl')
        else:
            st.success("No hay docentes pendientes de revisión.")

    with res_tabs[7]:
        st.markdown("### Descargar Resultados")

        # Estilizar el Excel principal
        styled_excel = _style_excel_workbook(excel_bytes)

        col_d1, col_d2, col_d3 = st.columns(3)
        with col_d1:
            st.download_button(
                "📥 Excel Formateado",
                data=styled_excel,
                file_name=f"BRP_ANUAL_{anio}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_anual_xlsx",
            )
        with col_d2:
            try:
                # PDF resumen con las tablas principales
                df_pdf = pd.DataFrame(table_rows)
                df_pdf = append_totals_row(df_pdf, label_col='MES', label='TOTAL')
                pdf_resumen = generate_pdf_from_df(df_pdf, title=f'Resumen Anual BRP {anio}')
                st.download_button(
                    "📥 PDF Resumen",
                    data=pdf_resumen,
                    file_name=f"BRP_RESUMEN_{anio}.pdf",
                    mime="application/pdf",
                    key="dl_anual_pdf",
                )
            except Exception:
                st.caption("PDF no disponible")
        with col_d3:
            # CSV del detalle completo
            try:
                df_csv = _read_detalle_brp(excel_bytes)
                csv_data = df_csv.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    "📥 CSV Detalle",
                    data=csv_data,
                    file_name=f"BRP_DETALLE_{anio}.csv",
                    mime="text/csv",
                    key="dl_anual_csv",
                )
            except Exception:
                st.caption("CSV no disponible")

        st.caption("El Excel incluye todas las hojas: DETALLE_BRP, POR_RBD, REVISAR, DETALLE_SEP, DETALLE_PIE")


def tab_lote_anual():
    """Pestaña de procesamiento anual por lotes."""

    with st.expander("📖 ¿Cómo usar el Lote Anual?", expanded=False):
        show_tutorial([
            ("Prepara los archivos", "Opción A: SEP y PIE por mes + web sostenedor. Opción B: un archivo anual consolidado (con columnas Periodo y Tipo_de_Contrato) + web sostenedor. Opcionalmente, agrega un archivo de horas por subvención (con columnas Mes, Rut, SEP, PIE, SN) para distribución real."),
            ("Nombra los archivos", "Los archivos se detectan por nombre. Ejemplo: 'web_sostenedor.xlsx', 'SEP ENERO 2026.xlsx'. Un archivo anual consolidado se detecta automáticamente."),
            ("Sube todos de golpe", "Arrastra todos los archivos al uploader. Puedes subir varios años de golpe (ej: 24 archivos de 8 años) — el sistema detecta el año de cada archivo por su nombre."),
            ("Revisa la grilla", "Modo 1 año: verifica que cada mes tenga SEP, PIE y WEB. Modo multi-año: revisa la grilla de años y sus archivos."),
            ("Procesa", "Haz clic en el botón y espera. En modo multi-año se procesa cada año secuencialmente y se genera un ZIP con todos los Excels."),
        ])

    st.markdown("---")

    # Año (solo relevante para modo 1 año)
    col_anio, col_spacer = st.columns([1, 3])
    with col_anio:
        anio = st.number_input("Año (modo 1 año)", min_value=2015, max_value=2030, value=2026, key="anual_anio")

    st.markdown("")
    st.markdown("##### 📁 Archivos del Año")
    st.caption(
        "Sube todos los archivos. Se detectan automáticamente por nombre "
        "(web*, sep*, *pie*/*sn*, *eib*) y mes (enero, febrero, etc.). "
        "También acepta un archivo anual consolidado con columnas Periodo y Tipo_de_Contrato. "
        "Si los nombres incluyen años distintos (ej: 2017, 2018...) se activa el modo multi-año."
    )

    archivos = st.file_uploader(
        "Arrastra o selecciona los archivos (Excel o CSV)",
        type=['xlsx', 'xls', 'csv'],
        accept_multiple_files=True,
        key="anual_files"
    )

    if archivos:
        # Guardar archivos temporales y clasificar (preserva extensión original)
        tmp_paths = []
        file_tuples = []
        for f in archivos:
            original_ext = Path(f.name).suffix or '.xlsx'
            tmp = tempfile.NamedTemporaryFile(suffix=original_ext, delete=False)
            tmp.write(f.getvalue())
            tmp.close()
            p = Path(tmp.name)
            tmp_paths.append(p)
            file_tuples.append((f.name, p))

        # --- Detectar si hay múltiples años ---
        detected_years = {}
        for fname, fpath in file_tuples:
            yr = detect_year_from_filename(fname)
            if yr:
                detected_years.setdefault(yr, []).append((fname, fpath))

        unique_years = sorted(detected_years.keys())
        is_multi_year = len(unique_years) > 1

        if is_multi_year:
            _tab_lote_anual_multi(file_tuples, tmp_paths, detected_years, unique_years)
        else:
            _tab_lote_anual_single(file_tuples, tmp_paths, anio)


def _tab_lote_anual_single(file_tuples, tmp_paths, anio):
    """Modo un solo año — flujo original de lote anual."""
    processor = AnualBatchProcessor()
    monthly = processor.classify_files(file_tuples)

    # Indicar si hay web compartido
    shared_web = getattr(processor, '_shared_web', None)
    if shared_web:
        st.info(
            f"📋 **Web sostenedor compartido:** {shared_web[0]} "
            f"(se filtrará automáticamente por mes)"
        )

    # Indicar si se detectó archivo de horas reales
    horas_file = getattr(processor, '_horas_file', None)
    if horas_file:
        st.info(
            f"📋 **Archivo de horas por subvención detectado:** {horas_file[0]} "
            f"(distribución real SEP/PIE/SN)"
        )

    # Indicar si se detectó archivo anual consolidado
    anual_file = getattr(processor, '_anual_file', None)
    if anual_file:
        st.info(
            f"📊 **Archivo anual consolidado detectado:** {anual_file[0]} — "
            f"se dividió automáticamente en archivos SEP/PIE sintéticos por mes."
        )

    # Mostrar grilla de detección
    st.markdown("##### Detección de archivos")
    grid_data = []
    all_months = [f"{i:02d}" for i in range(1, 13)]
    for m in all_months:
        ms = monthly.get(m)
        web_name = ''
        if ms and ms.web:
            web_name = ms.web[0]
            if shared_web and ms.web[0] == shared_web[0]:
                web_name = f"{web_name} (compartido)"
        elif shared_web:
            web_name = f"{shared_web[0]} (compartido)"

        sep_name = ''
        pie_name = ''
        if ms and ms.sep:
            sep_name = ms.sep[0]
            if ms.pre_processed:
                sep_name = f"{sep_name} (auto)"
        if ms and ms.pie:
            pie_name = ms.pie[0]
            if ms.pre_processed:
                pie_name = f"{pie_name} (auto)"

        row = {
            'Mes': MESES_NUM_TO_NAME.get(m, m),
            'SEP': sep_name,
            'PIE': pie_name,
            'WEB': web_name,
            'EIB': ms.eib[0] if ms and ms.eib else '(opcional)',
        }
        grid_data.append(row)

    df_grid = pd.DataFrame(grid_data)
    st.dataframe(df_grid, width='stretch', hide_index=True)

    # Validar
    errors = processor.validate_monthly_sets(monthly)
    meses_listos = len(monthly) - len(errors)

    if errors:
        st.warning(f"Meses incompletos ({len(errors)}):")
        for err in errors:
            st.markdown(f"- {err}")

    if not monthly:
        st.error("No se detectaron meses válidos.")
        # Diagnóstico detallado
        detected_items = []
        if shared_web:
            detected_items.append(f"- **Web sostenedor compartido:** `{shared_web[0]}`")
        if horas_file:
            detected_items.append(f"- **Archivo de horas:** `{horas_file[0]}`")
        if anual_file:
            detected_items.append(f"- **Archivo anual consolidado:** `{anual_file[0]}`")

        if detected_items:
            st.info("**Archivos detectados (pero no generan meses):**\n" + "\n".join(detected_items))

        # Analizar cada archivo subido para mostrar columnas encontradas
        diag_lines = []
        for fname, fpath in file_tuples:
            ftype = detect_file_type(fname)
            fmonth = detect_month_from_filename(fname)
            # Leer columnas del archivo
            try:
                ext = fpath.suffix.lower()
                if ext == '.csv':
                    df_peek = pd.read_csv(str(fpath), nrows=3, encoding='latin-1')
                else:
                    df_peek = pd.read_excel(str(fpath), nrows=3, engine='openpyxl')
                cols_found = list(df_peek.columns)
                cols_str = ", ".join(str(c) for c in cols_found[:15])
                if len(cols_found) > 15:
                    cols_str += f" ... (+{len(cols_found)-15} más)"
            except Exception:
                cols_str = "(no se pudo leer)"

            status_parts = []
            if ftype:
                status_parts.append(f"tipo=`{ftype}`")
            else:
                status_parts.append("tipo=**no detectado**")
            if fmonth:
                status_parts.append(f"mes=`{MESES_NUM_TO_NAME.get(fmonth, fmonth)}`")
            else:
                status_parts.append("mes=**no detectado**")

            diag_lines.append(
                f"**`{fname}`**: {', '.join(status_parts)}\n"
                f"  Columnas: {cols_str}"
            )

        if diag_lines:
            st.warning(
                "**Diagnóstico por archivo:**\n\n" + "\n\n".join(diag_lines)
            )

        st.info(
            "**Para generar meses se necesita al menos una de estas opciones:**\n"
            "1. Archivos SEP + PIE por mes (ej: `sep_enero.xlsx`, `sn_enero.xlsx`)\n"
            "2. Un archivo anual consolidado con columnas `Periodo` y `Tipo_de_Contrato`\n\n"
            "Los nombres deben contener el tipo (`sep`, `sn`/`pie`, `web`, `eib`) "
            "y el mes (`enero`, `ene`, etc.)."
        )
        _cleanup_temp_files(*tmp_paths)
        return

    st.markdown("")
    col_info1, col_info2 = st.columns(2)
    with col_info1:
        st.metric("Meses detectados", len(monthly))
    with col_info2:
        st.metric("Meses listos", meses_listos)

    # Solo procesar meses completos
    meses_completos = {
        k: v for k, v in monthly.items()
        if v.sep and v.pie and v.web
    }

    if not meses_completos:
        st.error("No hay meses con archivos completos (SEP + PIE + WEB).")
        _cleanup_temp_files(*tmp_paths)
        return

    st.markdown("")
    if st.button(
        f"▶️  PROCESAR {len(meses_completos)} MESES",
        key="btn_anual",
        width='stretch'
    ):
        progress = st.progress(0)
        status = st.empty()

        def callback(val, msg):
            progress.progress(min(val, 100) / 100)
            status.markdown(f"**{msg}**")

        out_path = None
        try:
            out_tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            out_path = Path(out_tmp.name)
            out_tmp.close()

            stats = processor.process_all(meses_completos, out_path, callback)

            # Leer bytes antes del cleanup
            with open(out_path, 'rb') as f:
                excel_bytes = f.read()

            st.session_state['lote_result'] = {
                'stats': stats,
                'excel_bytes': excel_bytes,
                'anio': anio,
            }
        except Exception as e:
            st.error(f"Error en procesamiento anual: {e}")
            import traceback
            with st.expander("Ver detalles técnicos"):
                st.code(traceback.format_exc())

        finally:
            _cleanup_temp_files(*tmp_paths, out_path)

    # Mostrar resultados cacheados
    if 'lote_result' in st.session_state:
        cached = st.session_state['lote_result']
        _display_lote_anual_results(
            cached['stats'], cached['excel_bytes'], cached['anio']
        )


def _tab_lote_anual_multi(file_tuples, tmp_paths, detected_years, unique_years):
    """Modo multi-año: procesa varios años secuencialmente y genera ZIP."""
    import gc
    import zipfile

    st.info(
        f"**Modo multi-año detectado:** {len(unique_years)} años "
        f"({unique_years[0]}–{unique_years[-1]}), "
        f"{len(file_tuples)} archivos en total."
    )

    # Archivos sin año detectado
    unmatched = [(fn, fp) for fn, fp in file_tuples
                 if detect_year_from_filename(fn) is None]

    # Grilla de detección por año
    st.markdown("##### Archivos detectados por año")
    grid_rows = []
    for yr in unique_years:
        yr_files = detected_years[yr]
        anual_count = 0
        web_count = 0
        horas_count = 0
        other_count = 0
        for fn, _fp in yr_files:
            ft = detect_file_type(fn)
            fn_lower = fn.lower()
            if ft == 'web':
                web_count += 1
            elif 'hora' in fn_lower or 'contrato' in fn_lower:
                horas_count += 1
            elif 'anual' in fn_lower or 'consolidado' in fn_lower:
                anual_count += 1
            elif ft in ('sep', 'pie', 'eib'):
                other_count += 1
            else:
                other_count += 1
        grid_rows.append({
            'Ano': yr,
            'Archivos': len(yr_files),
            'Anual/Consolidado': anual_count if anual_count else '-',
            'Web Sostenedor': web_count if web_count else '-',
            'Horas x Contrato': horas_count if horas_count else '-',
            'Otros (SEP/PIE/etc)': other_count if other_count else '-',
        })

    df_year_grid = pd.DataFrame(grid_rows)
    st.dataframe(df_year_grid, width='stretch', hide_index=True)

    if unmatched:
        st.warning(
            f"**{len(unmatched)} archivo(s) sin año detectado** "
            f"(se ignorarán en modo multi-año): "
            + ", ".join(f"`{fn}`" for fn, _ in unmatched)
        )

    # Detalle expandible por año
    for yr in unique_years:
        yr_files = detected_years[yr]
        with st.expander(f"Archivos del {yr} ({len(yr_files)})"):
            for fn, _fp in yr_files:
                ft = detect_file_type(fn) or "?"
                st.markdown(f"- `{fn}` — tipo: **{ft}**")

    st.markdown("")
    total_archivos_con_anio = sum(len(v) for v in detected_years.values())
    if st.button(
        f"▶️  PROCESAR {len(unique_years)} ANOS ({total_archivos_con_anio} archivos)",
        key="btn_anual_multi",
        width='stretch'
    ):
        progress_global = st.progress(0)
        status_global = st.empty()
        results = {}  # {año: styled_excel_bytes}
        all_stats = {}  # {año: stats}
        year_errors = {}

        for idx, yr in enumerate(unique_years):
            yr_files = detected_years[yr]
            status_global.markdown(f"**Procesando año {yr}** ({idx+1}/{len(unique_years)})...")
            base_progress = idx / len(unique_years)

            # Crear processor para este año
            processor = AnualBatchProcessor()
            monthly = processor.classify_files(yr_files)

            if not monthly:
                year_errors[yr] = "No se detectaron meses válidos."
                continue

            meses_completos = {
                k: v for k, v in monthly.items()
                if v.sep and v.pie and v.web
            }

            if not meses_completos:
                year_errors[yr] = "No hay meses con archivos completos (SEP + PIE + WEB)."
                continue

            out_path = None
            try:
                out_tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                out_path = Path(out_tmp.name)
                out_tmp.close()

                def make_callback(base_pct, year_weight):
                    def callback(val, msg):
                        combined = base_pct + (val / 100) * year_weight
                        progress_global.progress(min(combined, 1.0))
                        status_global.markdown(f"**Ano {yr}:** {msg}")
                    return callback

                year_weight = 1.0 / len(unique_years)
                stats = processor.process_all(
                    meses_completos, out_path,
                    make_callback(base_progress, year_weight)
                )

                with open(out_path, 'rb') as f:
                    raw_bytes = f.read()

                styled_bytes = _style_excel_workbook(raw_bytes)
                results[yr] = styled_bytes
                all_stats[yr] = stats

            except Exception as e:
                year_errors[yr] = str(e)
            finally:
                if out_path:
                    _cleanup_temp_files(out_path)
                # Liberar memoria entre años
                del processor
                gc.collect()

        progress_global.progress(1.0)
        status_global.markdown("**Procesamiento multi-año completado.**")

        # Limpiar temp files originales
        _cleanup_temp_files(*tmp_paths)

        # Guardar en session_state
        if results:
            st.session_state['lote_multi_result'] = {
                'results': results,
                'all_stats': all_stats,
                'year_errors': year_errors,
                'unique_years': unique_years,
            }

    # Mostrar resultados cacheados
    if 'lote_multi_result' in st.session_state:
        import zipfile
        cached = st.session_state['lote_multi_result']
        results = cached['results']
        all_stats = cached['all_stats']
        year_errors = cached['year_errors']

        st.markdown("---")
        st.markdown("### Resultados Multi-Ano")

        # Resumen por año
        if year_errors:
            for yr, err in sorted(year_errors.items()):
                st.error(f"**{yr}:** {err}")

        if results:
            success_box(
                f"**{len(results)} de {len(cached['unique_years'])} anos** "
                f"procesados exitosamente"
            )

            # Tabla resumen consolidada
            resumen_rows = []
            for yr in sorted(results.keys()):
                stats = all_stats.get(yr, {})
                summaries = [s for s in stats.get('summaries', []) if 'ERROR' not in s]
                brp_total = sum(
                    s.get('BRP_SEP', 0) + s.get('BRP_PIE', 0) + s.get('BRP_NORMAL', 0)
                    for s in summaries
                )
                docentes = max((s.get('DOCENTES_BRP', 0) for s in summaries), default=0)
                meses_ok = stats.get('meses_procesados', 0)
                resumen_rows.append({
                    'Ano': yr,
                    'Meses': meses_ok,
                    'Docentes (max)': docentes,
                    'BRP Total': int(brp_total),
                })

            if resumen_rows:
                df_resumen = pd.DataFrame(resumen_rows)
                # Agregar fila TOTAL
                df_resumen = append_totals_row(df_resumen, 'Ano', 'TOTAL')
                money_fmt = format_money_cols(df_resumen)
                st.dataframe(
                    df_resumen.style.format(money_fmt),
                    width='stretch', hide_index=True
                )

            # ZIP con todos los años
            buf = BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for yr, excel_bytes in sorted(results.items()):
                    zf.writestr(f"BRP_ANUAL_{yr}.xlsx", excel_bytes)
            zip_bytes = buf.getvalue()

            st.download_button(
                f"Descargar ZIP ({len(results)} anos)",
                data=zip_bytes,
                file_name=f"BRP_ANUAL_{min(results)}_{max(results)}.zip",
                mime="application/zip",
                key="btn_download_multi_zip",
                width='stretch',
            )

            # Descargas individuales
            st.markdown("##### Descargas individuales")
            cols = st.columns(min(len(results), 4))
            for i, (yr, excel_bytes) in enumerate(sorted(results.items())):
                with cols[i % len(cols)]:
                    st.download_button(
                        f"{yr}",
                        data=excel_bytes,
                        file_name=f"BRP_ANUAL_{yr}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"btn_download_yr_{yr}",
                    )

            # Detalle expandible por año
            for yr in sorted(results.keys()):
                stats = all_stats.get(yr, {})
                with st.expander(f"Detalle {yr}"):
                    _display_lote_anual_results(
                        stats, results[yr], yr
                    )




# ============================================================================
# HORAS POR CONTRATO
# ============================================================================

def _detect_mes_from_rem(filename: str) -> int | None:
    """Detecta el mes de un nombre de archivo REM.

    Patrones reconocidos:
        - 'REM 01 2026.xlsx' → 1
        - '2017-1.xlsx' → 1
        - '2026-12.csv' → 12
        - 'REM ENERO 2026.xlsx' → 1  (fallback por nombre de mes)
    """
    # Patrón REM + número
    m = re.search(r'REM\s*(\d{1,2})', filename, re.IGNORECASE)
    if m:
        num = int(m.group(1))
        if 1 <= num <= 12:
            return num
    # Patrón YYYY-M o YYYY-MM (ej: "2017-1", "2026-12")
    m = re.search(r'(\d{4})-(\d{1,2})', filename)
    if m:
        num = int(m.group(2))
        if 1 <= num <= 12:
            return num
    # Fallback: buscar mes por nombre (enero, febrero, etc.)
    detected = detect_month_from_filename(filename)
    if detected:
        return int(detected)
    return None


def _classify_tipocontrato(tipo: str) -> str:
    """Clasifica un tipo de contrato en SEP/PIE/EIB/SN."""
    if not tipo or not isinstance(tipo, str):
        return 'SN'
    t = str(tipo).upper().strip()
    if 'SEP' in t:
        return 'SEP'
    if 'PIE' in t:
        return 'PIE'
    if 'EIB' in t:
        return 'EIB'
    return 'SN'


def _process_rem_files(uploaded_files: list) -> tuple:
    """Procesa archivos REM y genera consolidado de horas por subvención.

    Returns:
        (df_resultado, alertas): DataFrame consolidado y lista de alertas.
    """
    alertas = []
    all_rows = []
    meses_vistos = {}

    for uf in uploaded_files:
        filename = uf.name
        mes_num = _detect_mes_from_rem(filename)

        if mes_num is None:
            alertas.append({
                'TIPO': 'ERROR', 'NIVEL': 'error',
                'ARCHIVO': filename,
                'MENSAJE': f'No se pudo detectar el mes del archivo "{filename}"',
            })
            continue

        mes_name = MESES_NUM_TO_NAME.get(f'{mes_num:02d}', f'Mes {mes_num}')

        # Detectar mes duplicado
        if mes_num in meses_vistos:
            alertas.append({
                'TIPO': 'ERROR', 'NIVEL': 'error',
                'ARCHIVO': filename,
                'MENSAJE': f'Mes {mes_name} duplicado: ya cargado desde "{meses_vistos[mes_num]}"',
            })
            continue
        meses_vistos[mes_num] = filename

        # Leer archivo
        try:
            ext = Path(filename).suffix.lower()
            if ext == '.csv':
                try:
                    df = pd.read_csv(uf, encoding='utf-8')
                except UnicodeDecodeError:
                    uf.seek(0)
                    df = pd.read_csv(uf, encoding='latin-1')
            else:
                df = pd.read_excel(uf, sheet_name=0, engine='openpyxl')
        except Exception as e:
            alertas.append({
                'TIPO': 'ERROR', 'NIVEL': 'error',
                'ARCHIVO': filename,
                'MENSAJE': f'Error al leer archivo: {e}',
            })
            continue

        df.columns = df.columns.str.strip()

        # Detectar columnas
        col_rut = _find_col(df, 'rut')
        col_nombre = _find_col(df, 'nombre')
        col_tipo = _find_col(df, 'tipocontrato', 'tipo_contrato', 'tipo contrato', 'tipo_de_contrato')
        col_escalafon = _find_col(df, 'escalafon', 'escalafón')
        col_depto = _find_col(df, 'departamento', 'ubicacion', 'ubicación')
        col_jornada = _find_col(df, 'jornada', 'horas')

        if not col_rut:
            alertas.append({
                'TIPO': 'ERROR', 'NIVEL': 'error',
                'ARCHIVO': filename,
                'MENSAJE': 'No se encontró columna RUT',
            })
            continue
        if not col_jornada:
            alertas.append({
                'TIPO': 'ERROR', 'NIVEL': 'error',
                'ARCHIVO': filename,
                'MENSAJE': 'No se encontró columna de jornada/horas',
            })
            continue

        # Procesar cada fila
        for _, row in df.iterrows():
            rut_raw = row.get(col_rut)
            if pd.isna(rut_raw) or not str(rut_raw).strip():
                continue

            rut = format_rut(str(rut_raw))
            nombre = str(row.get(col_nombre, '')).strip() if col_nombre else ''
            if nombre == 'nan':
                nombre = ''

            tipo_raw = str(row.get(col_tipo, '')).strip() if col_tipo else ''
            subvencion = _classify_tipocontrato(tipo_raw)

            if col_tipo and not tipo_raw:
                alertas.append({
                    'TIPO': 'WARNING', 'NIVEL': 'warning',
                    'ARCHIVO': filename,
                    'MENSAJE': f'Tipo contrato vacío para RUT {rut} — asignado a SN',
                })

            jornada = 0
            try:
                jornada = float(row.get(col_jornada, 0) or 0)
            except (ValueError, TypeError):
                pass

            # Resolver escuela/RBD
            depto = str(row.get(col_depto, '')).strip() if col_depto else ''
            if depto == 'nan':
                depto = ''
            escuela = depto  # Nombre original tal cual viene del REM
            rbd_gasto = ''
            if depto:
                match = match_ubicacion(depto)
                if match:
                    _esc_name, rbd_dv = match
                    if rbd_dv == 'DEM':
                        rbd_gasto = 'AC'  # Administración Central
                    else:
                        # RBD sin dígito verificador: "6718-0" → "6718"
                        rbd_gasto = rbd_dv.split('-')[0] if '-' in rbd_dv else rbd_dv
                else:
                    rbd_gasto = '?'
                    alertas.append({
                        'TIPO': 'WARNING', 'NIVEL': 'warning',
                        'ARCHIVO': filename,
                        'MENSAJE': f'Escuela sin match RBD: "{depto}" (RUT {rut})',
                    })

            cargo = str(row.get(col_escalafon, '')).strip() if col_escalafon else ''
            if cargo == 'nan':
                cargo = ''

            all_rows.append({
                'MES': mes_name,
                'MES_NUM': mes_num,
                'RUT': rut,
                'NOMBRE': nombre,
                'ESCUELA': escuela,
                'RBD GASTO SN': rbd_gasto,
                'SUBVENCION': subvencion,
                'JORNADA': jornada,
                'CARGO': cargo,
            })

    if not all_rows:
        return pd.DataFrame(), alertas

    df_raw = pd.DataFrame(all_rows)

    # Agrupar por RUT + ESCUELA + MES → sumar jornada por subvención (pivot)
    df_pivot = df_raw.pivot_table(
        index=['MES', 'MES_NUM', 'RUT', 'NOMBRE', 'ESCUELA', 'RBD GASTO SN', 'CARGO'],
        columns='SUBVENCION',
        values='JORNADA',
        aggfunc='sum',
        fill_value=0,
    ).reset_index()

    # Asegurar que las columnas existan
    for col in ['SEP', 'PIE', 'SN', 'EIB']:
        if col not in df_pivot.columns:
            df_pivot[col] = 0

    df_pivot['TOTAL HORAS'] = df_pivot['SEP'] + df_pivot['PIE'] + df_pivot['SN'] + df_pivot['EIB']
    df_pivot = df_pivot.sort_values(['MES_NUM', 'RUT', 'ESCUELA'])

    # Renombrar para compatibilidad con formato "horas por contrato"
    # y con lote anual (_is_horas_file busca: Mes, Rut, SEP, PIE, SN)
    df_pivot = df_pivot.rename(columns={
        'MES_NUM': 'Mes', 'RUT': 'Rut', 'NOMBRE': 'Nombre',
    })
    # Quitar MES (nombre texto) — Mes (número) es suficiente
    if 'MES' in df_pivot.columns:
        df_pivot = df_pivot.drop(columns=['MES'])

    cols_order = ['Mes', 'Rut', 'Nombre', 'ESCUELA', 'RBD GASTO SN', 'SEP', 'PIE', 'SN', 'EIB', 'TOTAL HORAS', 'CARGO']
    cols_order = [c for c in cols_order if c in df_pivot.columns]
    df_result = df_pivot[cols_order]

    # Alertas adicionales
    # Horas > 44 por persona/mes
    personas_mes = df_result.groupby(['Mes', 'Rut', 'Nombre']).agg({'TOTAL HORAS': 'sum'}).reset_index()
    exceden = personas_mes[personas_mes['TOTAL HORAS'] > 44]
    for _, row in exceden.iterrows():
        mes_name = MESES_NUM_TO_NAME.get(f'{int(row["Mes"]):02d}', f'Mes {int(row["Mes"])}')
        alertas.append({
            'TIPO': 'WARNING', 'NIVEL': 'warning',
            'ARCHIVO': '',
            'MENSAJE': f'{row["Nombre"]} ({row["Rut"]}) excede 44 hrs en {mes_name}: {row["TOTAL HORAS"]:.0f} hrs',
        })

    # Horas negativas en cualquier subvención
    for sub_col in ['SEP', 'PIE', 'SN', 'EIB']:
        if sub_col in df_result.columns:
            neg = df_result[df_result[sub_col] < 0]
            for _, row in neg.iterrows():
                mes_name = MESES_NUM_TO_NAME.get(f'{int(row["Mes"]):02d}', f'Mes {int(row["Mes"])}')
                alertas.append({
                    'TIPO': 'WARNING', 'NIVEL': 'warning',
                    'ARCHIVO': '',
                    'MENSAJE': f'{row["Nombre"]} ({row["Rut"]}) tiene {sub_col} negativo en {mes_name}: {row[sub_col]:.0f} hrs',
                })

    # Docente desaparece entre meses
    meses_procesados = sorted(df_result['Mes'].unique())
    if len(meses_procesados) > 1:
        for i in range(len(meses_procesados) - 1):
            m1, m2 = meses_procesados[i], meses_procesados[i + 1]
            ruts_m1 = set(df_result[df_result['Mes'] == m1]['Rut'].unique())
            ruts_m2 = set(df_result[df_result['Mes'] == m2]['Rut'].unique())
            desaparecidos = ruts_m1 - ruts_m2
            mes1_name = MESES_NUM_TO_NAME.get(f'{m1:02d}', f'Mes {m1}')
            mes2_name = MESES_NUM_TO_NAME.get(f'{m2:02d}', f'Mes {m2}')
            for rut in desaparecidos:
                nombre = df_result[df_result['Rut'] == rut]['Nombre'].iloc[0]
                alertas.append({
                    'TIPO': 'INFO', 'NIVEL': 'info',
                    'ARCHIVO': '',
                    'MENSAJE': f'{nombre} ({rut}) presente en {mes1_name} pero no en {mes2_name}',
                })

    return df_result, alertas


@st.fragment
def _render_horas_contrato_results(df_result, alertas):
    """Renderiza los resultados del procesamiento de horas por contrato."""
    # Métricas
    mc1, mc2, mc3, mc4 = st.columns(4)
    with mc1:
        st.metric("Personas", df_result['Rut'].nunique())
    with mc2:
        st.metric("Registros", len(df_result))
    with mc3:
        st.metric("Meses", df_result['Mes'].nunique())
    with mc4:
        total_horas = int(df_result['TOTAL HORAS'].sum())
        st.metric("Total Horas", f"{total_horas:,}".replace(',', '.'))

    st.markdown("---")

    # Filtros
    col_f1, col_f2, col_f3 = st.columns([1, 1, 2])
    with col_f1:
        meses_nums = sorted(df_result['Mes'].unique().tolist())
        meses_labels = [MESES_NUM_TO_NAME.get(f'{int(m):02d}', f'Mes {m}') for m in meses_nums]
        mes_sel = st.selectbox("Mes", ["Todos"] + meses_labels, key="hc_mes")
    with col_f2:
        escuelas = sorted(df_result['ESCUELA'].unique().tolist())
        esc_sel = st.selectbox("Escuela", ["Todas"] + escuelas, key="hc_esc")
    with col_f3:
        busqueda = st.text_input("Buscar por nombre o RUT", key="hc_buscar")

    df_f = df_result.copy()
    if mes_sel != "Todos":
        # Convertir label a número
        mes_idx = meses_labels.index(mes_sel)
        mes_num = meses_nums[mes_idx]
        df_f = df_f[df_f['Mes'] == mes_num]
    if esc_sel != "Todas":
        df_f = df_f[df_f['ESCUELA'] == esc_sel]
    if busqueda:
        b = busqueda.lower()
        mask = (
            df_f['Rut'].astype(str).str.lower().str.contains(b, na=False) |
            df_f['Nombre'].astype(str).str.lower().str.contains(b, na=False)
        )
        df_f = df_f[mask]

    # Tabla sin fila TOTAL
    st.dataframe(df_f, width='stretch', hide_index=True)

    # Totales fijos debajo de la tabla (siempre visibles sin scroll)
    t_sep = int(df_f['SEP'].sum()) if 'SEP' in df_f.columns else 0
    t_pie = int(df_f['PIE'].sum()) if 'PIE' in df_f.columns else 0
    t_sn = int(df_f['SN'].sum()) if 'SN' in df_f.columns else 0
    t_eib = int(df_f['EIB'].sum()) if 'EIB' in df_f.columns else 0
    t_total = int(df_f['TOTAL HORAS'].sum()) if 'TOTAL HORAS' in df_f.columns else 0
    t_reg = len(df_f)

    tc1, tc2, tc3, tc4, tc5, tc6 = st.columns(6)
    with tc1:
        st.metric("SEP", t_sep)
    with tc2:
        st.metric("PIE", t_pie)
    with tc3:
        st.metric("SN", t_sn)
    with tc4:
        st.metric("EIB", t_eib)
    with tc5:
        st.metric("TOTAL HORAS", f"{t_total:,}".replace(',', '.'))
    with tc6:
        st.metric("Registros", t_reg)

    # Descarga sin totales
    add_table_downloads(df_f, 'horas_por_contrato', 'hc_dl', excel_title=None)

    # Alertas
    if alertas:
        st.markdown("---")
        st.markdown("### Alertas y Errores")

        n_error = len([a for a in alertas if a['NIVEL'] == 'error'])
        n_warn = len([a for a in alertas if a['NIVEL'] == 'warning'])
        n_info = len([a for a in alertas if a['NIVEL'] == 'info'])

        ac1, ac2, ac3 = st.columns(3)
        with ac1:
            st.metric("Errores", n_error)
        with ac2:
            st.metric("Advertencias", n_warn)
        with ac3:
            st.metric("Info", n_info)

        df_alertas = pd.DataFrame(alertas)

        # Filtro por tipo
        tipos_alerta = sorted(df_alertas['TIPO'].unique().tolist())
        tipo_sel = st.multiselect("Filtrar por tipo", tipos_alerta, default=tipos_alerta, key="hc_alerta_tipo")
        df_alertas_f = df_alertas[df_alertas['TIPO'].isin(tipo_sel)]

        st.dataframe(df_alertas_f, width='stretch', hide_index=True)
        st.caption(f"Mostrando {len(df_alertas_f)} de {len(alertas)} alertas")


def tab_horas_contrato():
    """Pestaña de procesamiento de horas por contrato desde archivos REM."""

    with st.expander("📖 ¿Cómo usar Horas por Contrato?", expanded=False):
        show_tutorial([
            ("Sube archivos REM", "Sube los archivos REM mensuales (Excel o CSV). El nombre debe contener el mes, ej: 'REM 01 2026.xlsx'."),
            ("Revisa la detección", "Verifica que el mes se detectó correctamente para cada archivo."),
            ("Procesa", "Haz clic en Procesar. Se clasifican las horas por tipo de subvención (SEP/PIE/SN/EIB) según el tipo de contrato."),
        ])

    st.markdown("---")
    st.markdown("##### 📁 Archivos REM Mensuales")
    st.caption(
        "Sube los archivos REM de cada mes. Se detecta automáticamente el mes del nombre del archivo. "
        "Las horas se clasifican por tipo de subvención según el campo 'tipocontrato'."
    )

    uploaded = st.file_uploader(
        "Archivos REM",
        type=['xlsx', 'xls', 'csv'],
        accept_multiple_files=True,
        key="hc_uploader",
    )

    if not uploaded:
        st.info("Sube uno o más archivos REM para comenzar.")
        return

    # Grilla de validación
    st.markdown("##### Archivos detectados")
    grid_rows = []
    for uf in uploaded:
        mes_num = _detect_mes_from_rem(uf.name)
        mes_name = MESES_NUM_TO_NAME.get(f'{mes_num:02d}', '?') if mes_num else '?'
        grid_rows.append({
            'Archivo': uf.name,
            'Mes': mes_name,
            'Mes #': mes_num or '?',
        })
    df_grid = pd.DataFrame(grid_rows)
    st.dataframe(df_grid, width='stretch', hide_index=True)

    errores_deteccion = [r for r in grid_rows if r['Mes #'] == '?']
    if errores_deteccion:
        st.warning(f"{len(errores_deteccion)} archivo(s) sin mes detectado. Renombra los archivos incluyendo el mes (ej: 'REM 01 2026.xlsx').")

    # Botón procesar
    if st.button("Procesar Horas por Contrato", type="primary", key="hc_procesar"):
        with st.spinner("Procesando archivos REM..."):
            # Reset file position
            for uf in uploaded:
                uf.seek(0)
            df_result, alertas = _process_rem_files(uploaded)

        if df_result.empty:
            st.error("No se pudieron procesar los archivos.")
            if alertas:
                for a in alertas:
                    if a['NIVEL'] == 'error':
                        st.error(a['MENSAJE'])
                    else:
                        st.warning(a['MENSAJE'])
            return

        # Guardar en session_state
        st.session_state['hc_result'] = {
            'df': df_result,
            'alertas': alertas,
        }
        st.success(f"Procesados {df_result['Mes'].nunique()} meses, {df_result['Rut'].nunique()} personas, {len(df_result)} registros.")

    # Mostrar resultados cacheados
    if 'hc_result' in st.session_state:
        cached = st.session_state['hc_result']
        _render_horas_contrato_results(cached['df'], cached['alertas'])


# ============================================================================
# MAIN
# ============================================================================

def main():
    # Sidebar
    with st.sidebar:
        st.markdown("### 📊 RemuPro")
        st.caption(f"v{VERSION}")
        st.markdown("---")

        escuelas = get_rbd_map()
        st.markdown(f"**🏫 Establecimientos:** {len(escuelas)}")

        # Mostrar lista de escuelas
        with st.expander("Ver escuelas"):
            for rbd, nombre in sorted(escuelas.items(), key=lambda x: x[1]):
                st.markdown(f"**{rbd}** — {nombre}")

        st.markdown("---")
        st.markdown("**Archivos requeridos:**")
        st.markdown("""
        - `web*` → MINEDUC
        - `sep*` → SEP procesado
        - `sn*` / `*pie*` → PIE/Normal
        """)

        st.markdown("---")
        st.caption("DAEM = Subvención (municipio)")
        st.caption("CPEIP = Transferencia (ministerio)")

        # Acceso en red local
        st.markdown("---")
        try:
            local_ip = socket.gethostbyname(socket.gethostname())
            st.caption(f"Red local: http://{local_ip}:8501")
        except Exception:
            pass

        # Preferencias de pestañas visibles
        st.markdown("---")
        with st.expander("Pestañas visibles", expanded=False):
            show_brp = st.checkbox("Distribución BRP", value=False, key="pref_brp")
            show_dup = st.checkbox("Duplicados", value=False, key="pref_dup")

        # Charts dinámicos del último procesamiento
        st.markdown("---")
        show_sidebar_charts()

    show_header()

    # Tabs dinámicos según preferencias
    tab_names = ["⚡ Todo en Uno", "📊 SEP / PIE / EIB", "📅 Lote Anual", "⏱ Horas x Contrato"]
    tab_funcs = [tab_todo_en_uno, tab_sep_pie, tab_lote_anual, tab_horas_contrato]

    if st.session_state.get("pref_brp"):
        tab_names.append("💰 Distribución BRP")
        tab_funcs.append(tab_brp)
    if st.session_state.get("pref_dup"):
        tab_names.append("🔄 Duplicados")
        tab_funcs.append(tab_duplicados)

    tabs = st.tabs(tab_names)
    for tab, func in zip(tabs, tab_funcs):
        with tab:
            func()

    # Footer
    st.markdown(f"""
    <div class="app-footer">
        RemuPro v{VERSION} • Procesamiento de Remuneraciones Educativas
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
